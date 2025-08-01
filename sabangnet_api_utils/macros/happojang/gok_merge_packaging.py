"""G옥 주문 합포장 처리 모듈"""

from __future__ import annotations
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from utils.excels.excel_handler import ExcelHandler


# 설정 상수
MALL_NAME = "G·옥"

# 시트 분리 설정
ACCOUNT_MAPPING = {
    "OK,CL,BB": ["오케이마트", "클로버프", "베이지베이글"],
    "IY": ["아이예스"],
}

# 필수 생성 시트 목록 (항상 생성)
REQUIRED_SHEETS = list(ACCOUNT_MAPPING.keys()) 


class GokDataProcessor:
    """G옥 데이터 정리 처리 유틸리티"""
    MULTI_SEP_RE = re.compile(r"[\/;]")
    BRACKET_RE = re.compile(r"\[(.*?)\]")
    
    @staticmethod
    def clean_model_name(txt: str | None) -> str:
        """
        모델명 문자열 정리
        - '/' 또는 ';' → ' + '
        - ' 1개' 제거
        """
        if not txt:
            return ""
        txt = GokDataProcessor.MULTI_SEP_RE.sub(" + ", str(txt))
        return txt.replace(" 1개", "").strip()

    @staticmethod
    def extract_bracket_content(text: str | None) -> str:
        """[계정명] 형식에서 계정명만 추출"""
        if not text:
            return ""
        match = GokDataProcessor.BRACKET_RE.search(str(text))
        return match.group(1) if match else ""


def process_slash_values(ws: Worksheet) -> None:
    """V열의 슬래시(/) 구분 값 처리 - 첫 번째 유효 숫자만 사용"""
    for r in range(2, ws.max_row + 1):
        v_raw = str(ws[f"V{r}"].value or "").strip()
        if "/" in v_raw:
            nums = [
                int(n)
                for n in v_raw.split("/")
                if n.strip().isdigit() and int(n) != 0
            ]
            ws[f"V{r}"].value = 0 if not nums else nums[0]


def truncate_order_numbers(ws: Worksheet, max_length: int = 10) -> None:
    """E열 주문번호 자르기"""
    for r in range(2, ws.max_row + 1):
        cell = ws[f"E{r}"]
        cell.value = str(cell.value)[:max_length]
        cell.number_format = "General"


def clear_l_column(ws: Worksheet) -> None:
    """L열(12번째 열) 내용 비우기"""
    # L열은 12번째 열 (1-based index)
    l_col_idx = 12
    
    # L열이 존재하는지 확인
    if l_col_idx <= ws.max_column:
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=l_col_idx).value = None
    
    # 혹시 헤더명으로도 찾아서 처리 (보조)
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col_idx).value or "").strip().upper()
        if header == "L":
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col_idx).value = None
            break


class GokSheetManager:
    """G옥 시트 분리 및 자동화 로직 적용"""
    
    def __init__(self, ws: Worksheet, account_mapping: Dict[str, List[str]]):
        self.ws = ws
        self.account_mapping = account_mapping
        self.last_row = ws.max_row
        self.last_col = ws.max_column
        
        # 열 너비 저장
        self.col_widths = [
            ws.column_dimensions[get_column_letter(c)].width
            for c in range(1, self.last_col + 1)
        ]

    def get_rows_by_sheet(self) -> Dict[str, List[int]]:
        """시트별 행 번호 매핑 생성
        - 일반 시트: 계정별 데이터만
        - 자동화_합포장_시트: 모든 데이터 포함
        """
        rows_by_sheet = defaultdict(list)
        all_rows = list(range(2, self.last_row + 1))

        # 자동화_합포장_시트는 모든 데이터 포함
        rows_by_sheet["자동화_합포장_시트"] = all_rows
        
        # 나머지 시트는 계정별로 데이터 분리
        for r in all_rows:
            account = GokDataProcessor.extract_bracket_content(
                self.ws[f"B{r}"].value
            )
            for sheet_name, accounts in self.account_mapping.items():
                if sheet_name != "자동화_합포장_시트" and account in accounts:
                    rows_by_sheet[sheet_name].append(r)
                    
        return rows_by_sheet

    def create_empty_sheet(self, wb: Worksheet, sheet_name: str) -> Worksheet:
        """빈 시트 생성 (헤더와 열 너비만 복사)"""
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            
        new_ws = wb.create_sheet(sheet_name)
        
        # 헤더와 열 너비 복사
        for c in range(1, self.last_col + 1):
            new_ws.cell(row=1, column=c, 
                       value=self.ws.cell(row=1, column=c).value)
            new_ws.column_dimensions[get_column_letter(c)].width = self.col_widths[c - 1]
            
        return new_ws

    def copy_sheet_data(self, ws: Worksheet, row_indices: List[int]) -> None:
        """시트에 데이터 행 복사"""
        if not row_indices:
            return
            
        for idx, r in enumerate(row_indices, start=2):
            for c in range(1, self.last_col + 1):
                ws.cell(row=idx, column=c, 
                       value=self.ws.cell(row=r, column=c).value)
            ws[f"A{idx}"].value = "=ROW()-1"

    def apply_automation_logic(self, ws: Worksheet) -> None:
        """자동화 로직 적용"""
        # 1. 기본 서식 적용
        ex = ExcelHandler(ws)
        ex.set_basic_format()
        
        # 2. P열 슬래시(/) 금액 합산
        ex.sum_prow_with_slash()
        
        # 3. V열 슬래시(/) 정제 
        # TODO - 변경 필요
        process_slash_values(ws)
        
        # 4. F열 모델명 정리
        for r in range(2, ws.max_row + 1):
            ws[f"F{r}"].value = GokDataProcessor.clean_model_name(ws[f"F{r}"].value)
            
        # 5. E열 주문번호 처리
        truncate_order_numbers(ws)
        
        # 6. 정렬 및 순번
        ex.set_column_alignment()
        ex.set_row_number(ws)
        
        # 7. 문자열→숫자 변환
        ex.convert_numeric_strings(cols=("E", "F", "M", "Q", "W"))
        
        # 8. C→B 2단계 정렬
        ex.sort_by_columns([2, 3])
        
        # 9. D열 수식 설정
        # ex.autofill_d_column(formula="=O{row}+P{row}+V{row}")
        ex.calculate_d_column_values(first_col='O', second_col='P', third_col='V')

        # 10. 서식 초기화
        ex.clear_fills_from_second_row()
        ex.clear_borders()
        # F열 왼쪽정렬 
        for row in range(1, ws.max_row + 1):
            ws[f"F{row}"].alignment = Alignment(horizontal='left')
        clear_l_column(ws)

    def copy_to_new_sheet(self, 
                         wb: Worksheet, 
                         sheet_name: str, 
                         row_indices: List[int] = None) -> None:
        """지정된 행들로 새 시트 생성 (데이터가 없어도 빈 시트 생성)"""
        new_ws = self.create_empty_sheet(wb, sheet_name)
        if row_indices:
            self.copy_sheet_data(new_ws, row_indices)
            
        # 모든 시트에 자동화 로직 적용
        self.apply_automation_logic(new_ws)


def gok_merge_packaging(file_path: str) -> str:
    """G옥 주문 합포장 자동화 처리"""
    # Excel 파일 로드
    ex = ExcelHandler.from_file(file_path)
    
    # 첫 번째 시트(원본)에 자동화 로직 적용
    source_ws = ex.ws
    splitter = GokSheetManager(source_ws, ACCOUNT_MAPPING)
    splitter.apply_automation_logic(source_ws)
    
    # 계정별 시트 분리 및 필수 시트 생성
    splitter = GokSheetManager(source_ws, ACCOUNT_MAPPING)
    rows_by_sheet = splitter.get_rows_by_sheet()
    
    # 모든 필수 시트 생성 (데이터 유무와 무관)
    for sheet_name in REQUIRED_SHEETS:
        splitter.copy_to_new_sheet(
            ex.wb,
            sheet_name, 
            rows_by_sheet.get(sheet_name, [])
        )
    
    # 저장
    base_name = Path(file_path).stem  # 확장자 제거한 파일명
    output_path = ex.happojang_save_file(base_name=base_name)
    ex.wb.close()
    
    print(f"◼︎ [{MALL_NAME}] 합포장 자동화 완료!")
    return output_path


if __name__ == "__main__":
    test_file = "/Users/smith/Documents/github/OKMart/sabangnet_API/files/test-[기본양식]-합포장용.xlsx"
    gok_merge_packaging(test_file)
    print("모든 처리가 완료되었습니다!")