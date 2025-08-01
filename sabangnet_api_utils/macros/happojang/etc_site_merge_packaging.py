"""ê¸°íƒ€ì‚¬ì´íŠ¸ í•©í¬ì¥ ìë™í™” ëª¨ë“ˆ"""

from __future__ import annotations
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Set

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from utils.excels.excel_handler import ExcelHandler

import pandas as pd

# ì„¤ì • ìƒìˆ˜
MALL_NAME = "ê¸°íƒ€ì‚¬ì´íŠ¸"
RED_FONT = Font(color="FF0000", bold=True)
BLUE_FILL = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")

# ì‹œíŠ¸ ë¶„ë¦¬ ì„¤ì • 
ACCOUNT_MAPPING = {
    "OK": ["ì˜¤ì¼€ì´ë§ˆíŠ¸", "í´ë¡œë²„í”„"],
    "BB": ["ë² ì´ì§€ë² ì´ê¸€"],
    "IY": ["ì•„ì´ì˜ˆìŠ¤"]
}

# í•„ìˆ˜ ìƒì„± ì‹œíŠ¸ ëª©ë¡
REQUIRED_SHEETS = list(ACCOUNT_MAPPING.keys())

# ì‚¬ì´íŠ¸ ì„¤ì •
class ETCSiteConfig:
    # ë°°ì†¡ë¹„ ë¶„í•  ëŒ€ìƒ
    DELIVERY_SPLIT_SITES: Set[str] = {
        "ë¡¯ë°ì˜¨", "ë³´ë¦¬ë³´ë¦¬", "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´", "í†¡ìŠ¤í† ì–´"
    }
    
    # ë°°ì†¡ë¹„ ë¬´ë£Œ ì‚¬ì´íŠ¸
    FREE_DELIVERY_SITES: Set[str] = {"ì˜¤ëŠ˜ì˜ì§‘"}
    
    # ì£¼ë¬¸ë²ˆí˜¸ ê¸¸ì´ ì œí•œ 
    ORDER_NUMBER_LENGTHS: Dict[str, int] = {
        "YES24": 11,
        "CJì˜¨ìŠ¤íƒ€ì¼": 26,
        "GSSHOP": 21,
        "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´": 16,
        "ì—ì´ë¸”ë¦¬": 13,
        "ì˜¬ì›¨ì´ì¦ˆ": 36,  
        "ì¹´ì¹´ì˜¤ì„ ë¬¼í•˜ê¸°": 10,  
        "ì¹´ì¹´ì˜¤í†¡ìŠ¤í† ì–´": 10,  
        "ìœ„ë©”í”„": 13,
        "ì¸í„°íŒŒí¬": 12,
        "ì¿ íŒ¡": 13,
        "í‹°ëª¬": 12,
        "í•˜ì´ë§ˆíŠ¸": 12
    }
    
    # ìˆ«ì ë³€í™˜ ëŒ€ìƒ ì‚¬ì´íŠ¸
    NUMERIC_SITES: Set[str] = {
        "ì—ì´ë¸”ë¦¬", "ì˜¤ëŠ˜ì˜ì§‘", "ì¿ íŒ¡", "í…ë°”ì´í…", "NSí™ˆì‡¼í•‘", 
        "ê·¸ë¦½", "ë³´ë¦¬ë³´ë¦¬", "ì¹´ì¹´ì˜¤ì„ ë¬¼í•˜ê¸°", "í†¡ìŠ¤í† ì–´", "í† ìŠ¤"
    }

class ETCOrderUtils:
    """ì£¼ë¬¸ë²ˆí˜¸ ì²˜ë¦¬ ìœ í‹¸ë¦¬í‹°"""
    
    @staticmethod
    def clean_order_text(txt) -> str:
        """ì£¼ë¬¸ë²ˆí˜¸ ë¬¸ìì—´ ì •ë¦¬ (ì•ˆì „í•œ íƒ€ì… ë³€í™˜ í¬í•¨)"""
        if not txt:
            return ""
        # ì•ˆì „í•œ ë¬¸ìì—´ ë³€í™˜ (float, int, str ëª¨ë‘ ì²˜ë¦¬)
        txt = str(txt).replace(" 1ê°œ", "").strip()
        txt = txt.replace("/", " + ")
        return txt

    @staticmethod
    def extract_bracket_text(text: str | None) -> str:
        """[ê³„ì •ëª…] í˜•ì‹ì—ì„œ ê³„ì •ëª…ë§Œ ì¶”ì¶œ"""
        if not text:
            return ""
        match = re.search(r"\[(.*?)\]", str(text))
        return match.group(1) if match else ""

def process_order_numbers(ws: Worksheet) -> None:
    """
    ğŸ”„ ExcelHandler í›„ë³´
    ì‚¬ì´íŠ¸ë³„ ì£¼ë¬¸ë²ˆí˜¸ ì²˜ë¦¬
    """
    for row in range(2, ws.max_row + 1):
        site = str(ws[f"B{row}"].value or "")
        order_raw = str(ws[f"E{row}"].value or "")
        
        # ì‚¬ì´íŠ¸ë³„ ì£¼ë¬¸ë²ˆí˜¸ ê¸¸ì´ ì œí•œ ì ìš©
        for site_name, length in ETCSiteConfig.ORDER_NUMBER_LENGTHS.items():
            if site_name in site:
                ws[f"E{row}"].value = order_raw[:length]
                break
                
        # ì¿ íŒ¡ íŠ¹ìˆ˜ ì²˜ë¦¬
        if "ì¿ íŒ¡" in site and "/" in order_raw:
            slash_count = order_raw.count("/")
            pure_length = len(order_raw.replace("/", ""))
            each_len = pure_length // (slash_count + 1)
            ws[f"E{row}"].value = order_raw[:each_len]


def etc_site_merge_packaging(input_path: str) -> str:
    """ê¸°íƒ€ì‚¬ì´íŠ¸ ì£¼ë¬¸ í•©í¬ì¥ ìë™í™” ì²˜ë¦¬ (VBA ë§¤í¬ë¡œ 14ë‹¨ê³„ ì‹œíŠ¸ë¶„ë¦¬ í¬í•¨)"""
    # Excel íŒŒì¼ ë¡œë“œ
    ex = ExcelHandler.from_file(input_path)
    ws = ex.ws
    
    # ========== VBA ë§¤í¬ë¡œ ë‹¨ê³„: ì›ë³¸ ì‹œíŠ¸ ìë™í™” ì²˜ë¦¬ ==========
    # ì›ë³¸ ì‹œíŠ¸ì— ìë™í™” ë¡œì§ ì ìš© (í–‰ ë¶„í•  í¬í•¨)
    splitter = ETCSheetManager(ws, ACCOUNT_MAPPING)
    splitter.apply_automation_logic(ws)
    
    # ========== ìë™í™” ì‹œíŠ¸ ìƒì„± (ë§¨ ì•ì— ìœ„ì¹˜) ==========
    # ë§¤í¬ë¡œê°€ ì ìš©ëœ ì „ì²´ ì‹œíŠ¸ë¥¼ "ìë™í™”" ì´ë¦„ìœ¼ë¡œ ë§¨ ì•ì— ë³µì‚¬
    splitter.create_automation_sheet(ex.wb)
    
    # ========== VBA ë§¤í¬ë¡œ 14ë‹¨ê³„: ì‹œíŠ¸ë¶„ë¦¬ (OK, BB, IY) ==========
    # ìë™í™” ì²˜ë¦¬ê°€ ì™„ë£Œëœ ì›ë³¸ ì‹œíŠ¸ì—ì„œ ì‹œíŠ¸ë¶„ë¦¬ ìˆ˜í–‰
    rows_by_sheet = splitter.get_rows_by_sheet()
    
    # ëª¨ë“  í•„ìˆ˜ ì‹œíŠ¸ ìƒì„± (ë°ì´í„° ìœ ë¬´ì™€ ë¬´ê´€í•˜ê²Œ OK, BB, IY ì‹œíŠ¸ ìƒì„±)
    for sheet_name in REQUIRED_SHEETS:
        splitter.copy_to_new_sheet_simple(
            ex.wb,
            sheet_name, 
            rows_by_sheet.get(sheet_name, [])
        )
    
    # ì›ë³¸ ì‹œíŠ¸ ì‚­ì œ (ìë™í™” ì‹œíŠ¸ë¡œ ëŒ€ì²´ë˜ì—ˆìœ¼ë¯€ë¡œ)
    original_sheet_name = ws.title
    if original_sheet_name in ex.wb.sheetnames and original_sheet_name != "ìë™í™”":
        del ex.wb[original_sheet_name]
    
    # ì €ì¥
    base_name = Path(input_path).stem  # í™•ì¥ì ì œê±°í•œ íŒŒì¼ëª…
    output_path = ex.happojang_save_file(base_name=base_name)
    ex.wb.close()
    
    return output_path


class ETCDeliveryFeeHandler:
    """ì‚¬ì´íŠ¸ë³„ ë°°ì†¡ë¹„ ì²˜ë¦¬"""
    
    def __init__(self, ws: Worksheet):
        self.ws = ws
        
    def process_delivery_fee(self) -> None:
        """
        ğŸ”„ ExcelHandler í›„ë³´
        ì‚¬ì´íŠ¸ë³„ ë°°ì†¡ë¹„ ì²˜ë¦¬
        - ë¡¯ë°ì˜¨/ë³´ë¦¬ë³´ë¦¬ ë“±: ì£¼ë¬¸ ìˆ˜ë¡œ ë‚˜ëˆ„ê¸°
        - ì˜¤ëŠ˜ì˜ì§‘: ë¬´ë£Œë°°ì†¡
        - í† ìŠ¤: 3ë§Œì› ì´ìƒ ë¬´ë£Œ
        - ë¶„í• ëœ í–‰ì— ëŒ€í•œ ì¬ì²˜ë¦¬ í¬í•¨
        """
        for row in range(2, self.ws.max_row + 1):
            site = str(self.ws[f"B{row}"].value or "")
            order_text = str(self.ws[f"X{row}"].value or "")
            v_cell = self.ws[f"V{row}"]
            
            # Vì—´ ê°’ì´ "/" êµ¬ë¶„ì í˜•íƒœì¸ ê²½ìš° ì²« ë²ˆì§¸ ê°’ ì‚¬ìš©
            v_value = v_cell.value
            if v_value and "/" in str(v_value):
                v_val = float(str(v_value).split("/")[0].strip() or 0)
            else:
                v_val = float(v_value or 0) if v_value else 0
            
            # Uì—´ ê°’ë„ ë™ì¼í•˜ê²Œ ì²˜ë¦¬
            u_value = self.ws[f"U{row}"].value
            if u_value and "/" in str(u_value):
                u_val = float(str(u_value).split("/")[0].strip() or 0)
            else:
                u_val = float(u_value or 0) if u_value else 0

            # ë¶„í• ëœ í–‰ì—ì„œ Vì—´ì´ ë¹„ì–´ìˆëŠ” ê²½ìš° ë°°ì†¡ë¹„ ë¡œì§ ì¬ì ìš©
            if v_value is None or v_value == "":
                
                # ë°°ì†¡ë¹„ ë¶„í•  ëŒ€ìƒ
                if any(s in site for s in ETCSiteConfig.DELIVERY_SPLIT_SITES) and "/" in order_text:
                    count = len(order_text.split("/"))
                    # ê¸°ë³¸ ë°°ì†¡ë¹„ 3000ì› ì ìš© í›„ ë¶„í• 
                    if count > 0:
                        v_cell.value = round(3000 / count)
                        v_cell.font = RED_FONT

                # ë¬´ë£Œë°°ì†¡
                elif any(s in site for s in ETCSiteConfig.FREE_DELIVERY_SITES):
                    v_cell.value = 0
                    v_cell.font = RED_FONT

                # í† ìŠ¤ (3ë§Œì› ì´ìƒ ë¬´ë£Œ)
                elif "í† ìŠ¤" in site:
                    v_cell.value = 0 if u_val > 30000 else 3000
                    v_cell.font = RED_FONT
                    
            else:
                # Vì—´ì— ê°’ì´ ìˆëŠ” ê²½ìš° ê¸°ì¡´ ë¡œì§ ì ìš©
                
                # ë°°ì†¡ë¹„ ë¶„í•  ëŒ€ìƒ
                if any(s in site for s in ETCSiteConfig.DELIVERY_SPLIT_SITES) and "/" in order_text:
                    count = len(order_text.split("/"))
                    if v_val > 3000 and count > 0:
                        v_cell.value = round(v_val / count)
                        v_cell.font = RED_FONT

                # ë¬´ë£Œë°°ì†¡
                elif any(s in site for s in ETCSiteConfig.FREE_DELIVERY_SITES):
                    v_cell.value = 0
                    v_cell.font = RED_FONT

                # í† ìŠ¤ (3ë§Œì› ì´ìƒ ë¬´ë£Œ)
                elif "í† ìŠ¤" in site:
                    v_cell.value = 0 if u_val > 30000 else 3000
                    v_cell.font = RED_FONT


class ETCSpecialCaseHandler:
    """íŠ¹ìˆ˜ ì¼€ì´ìŠ¤ ì²˜ë¦¬"""
    
    def __init__(self, ws: Worksheet):
        self.ws = ws
        
    def process_kakao_jeju(self) -> None:
        """ì¹´ì¹´ì˜¤ + ì œì£¼ë„ ì£¼ë¬¸ ì²˜ë¦¬"""
        for row in range(2, self.ws.max_row + 1):
            site = str(self.ws[f"B{row}"].value or "")
            addr = str(self.ws[f"J{row}"].value or "")
            
            if "ì¹´ì¹´ì˜¤" in site and "ì œì£¼" in addr:
                # Fì—´ ì•ˆë‚´ë¬¸êµ¬ ì¶”ê°€ ë° ë°°ê²½ìƒ‰
                f_cell = self.ws[f"F{row}"]
                if "[3000ì› ì—°ë½í•´ì•¼í•¨]" not in str(f_cell.value):
                    f_cell.value = f"{f_cell.value} [3000ì› ì—°ë½í•´ì•¼í•¨]"
                f_cell.fill = BLUE_FILL
                
                # Jì—´ ë¹¨ê°„ìƒ‰ êµµê²Œ
                self.ws[f"J{row}"].font = RED_FONT
                
    def process_l_column(self) -> None:
        """Lì—´ ì‹ ìš©/ì°©ë¶ˆ ì²˜ë¦¬"""
        for row in range(2, self.ws.max_row + 1):
            val = str(self.ws[f"L{row}"].value or "")
            if val == "ì‹ ìš©":
                self.ws[f"L{row}"].value = ""
            elif val == "ì°©ë¶ˆ":
                self.ws[f"L{row}"].font = RED_FONT


class ETCSheetManager:
    """ê¸°íƒ€ì‚¬ì´íŠ¸ ì‹œíŠ¸ ë¶„ë¦¬ ë° ìë™í™” ë¡œì§ ì ìš© (VBA ë§¤í¬ë¡œ 14ë‹¨ê³„ êµ¬í˜„)"""
    
    def __init__(self, ws: Worksheet, account_mapping: Dict[str, List[str]]):
        self.ws = ws
        self.account_mapping = account_mapping
        self.last_row = ws.max_row
        self.last_col = ws.max_column
        
        self.col_widths = [
            ws.column_dimensions[get_column_letter(c)].width
            for c in range(1, self.last_col + 1)
        ]

    def get_rows_by_sheet(self) -> Dict[str, List[int]]:
        """
        ì‹œíŠ¸ë³„ í–‰ ë²ˆí˜¸ ë§¤í•‘ ìƒì„± (VBA ExtractBracketText ë¡œì§ êµ¬í˜„)
        ìë™í™” ì²˜ë¦¬ê°€ ì™„ë£Œëœ ì›ë³¸ ì‹œíŠ¸ì—ì„œ í˜„ì¬ ìƒíƒœë¥¼ ê¸°ì¤€ìœ¼ë¡œ ë§¤í•‘
        """
        rows_by_sheet = defaultdict(list)
        
        # í˜„ì¬ ì‹œíŠ¸ì˜ ì‹¤ì œ ìµœëŒ€ í–‰ ìˆ˜ë¥¼ ì‚¬ìš© (ë¶„í• ëœ í–‰ í¬í•¨)
        current_max_row = self.ws.max_row
        
        for r in range(2, current_max_row + 1):
            # Bì—´ì—ì„œ [ê³„ì •ëª…] ì¶”ì¶œ (VBA ExtractBracketTextì™€ ë™ì¼)
            account = ETCOrderUtils.extract_bracket_text(self.ws[f"B{r}"].value)
            
            # ê° ì‹œíŠ¸ì˜ ê³„ì • ëª©ë¡ê³¼ ì •í™•íˆ ë¹„êµ
            for sheet_name, accounts in self.account_mapping.items():
                if account in accounts:
                    rows_by_sheet[sheet_name].append(r)
                    break
                    
        return rows_by_sheet

    def create_empty_sheet(self, wb, sheet_name: str) -> Worksheet:
        """ë¹ˆ ì‹œíŠ¸ ìƒì„± (í—¤ë”ì™€ ì—´ ë„ˆë¹„ë§Œ ë³µì‚¬) - VBA ë§¤í¬ë¡œì™€ ë™ì¼"""
        # ê¸°ì¡´ ì‹œíŠ¸ ì‚­ì œ (VBA: On Error Resume Next)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            
        new_ws = wb.create_sheet(sheet_name)
        
        # ì œëª©í–‰ ë³µì‚¬ (VBA: sourceSheet.Rows(1).Copy destSheet.Rows(1))
        for c in range(1, self.last_col + 1):
            new_ws.cell(row=1, column=c, 
                       value=self.ws.cell(row=1, column=c).value)
            # ì—´ë„ˆë¹„ ë³µì‚¬ (VBA: destSheet.Columns(c).ColumnWidth = colWidth(c))
            new_ws.column_dimensions[get_column_letter(c)].width = self.col_widths[c - 1]
            
        return new_ws

    def copy_sheet_data(self, ws: Worksheet, row_indices: List[int]) -> None:
        """ì‹œíŠ¸ì— ë°ì´í„° í–‰ ë³µì‚¬ (VBA: sourceSheet.Rows(r).Copy Destination:=destSheet.Rows(targetRow))"""
        if not row_indices:
            return
            
        target_row = 2
        for r in row_indices:
            for c in range(1, self.last_col + 1):
                ws.cell(row=target_row, column=c, 
                       value=self.ws.cell(row=r, column=c).value)
            target_row += 1
        
        # Aì—´ ìˆœë²ˆ ì¬ë¶€ì—¬ (VBA: destSheet.Cells(r, "A").Value = r - 1)
        for r in range(2, target_row):
            ws[f"A{r}"].value = r - 1

    def copy_to_new_sheet_simple(self, 
                                wb, 
                                sheet_name: str, 
                                row_indices: List[int] = None) -> None:
        """
        VBAì™€ ë™ì¼í•œ ë°©ì‹ìœ¼ë¡œ ìƒˆ ì‹œíŠ¸ ìƒì„±
        - ì›ë³¸ ì‹œíŠ¸ì˜ ì²˜ë¦¬ëœ ë°ì´í„°ë¥¼ ê° ê³„ì •ëª…ë³„ë¡œ ë³µì‚¬ë§Œ ìˆ˜í–‰
        - ì¶”ê°€ ìë™í™” ë¡œì§ ì ìš©í•˜ì§€ ì•ŠìŒ (VBA ë§¤í¬ë¡œì™€ ë™ì¼)
        """
        new_ws = self.create_empty_sheet(wb, sheet_name)
        if row_indices:
            self.copy_sheet_data_simple(new_ws, row_indices)
    
    def copy_sheet_data_simple(self, ws: Worksheet, row_indices: List[int]) -> None:
        """
        VBAì™€ ë™ì¼í•œ ë°©ì‹ìœ¼ë¡œ ì‹œíŠ¸ì— ë°ì´í„° í–‰ ë³µì‚¬
        - ì›ë³¸ ì‹œíŠ¸ì˜ í˜„ì¬ ìƒíƒœ(ë¶„í• ëœ í–‰ í¬í•¨)ë¥¼ ê·¸ëŒ€ë¡œ ë³µì‚¬
        - Aì—´ ìˆœë²ˆë§Œ ì¬ë¶€ì—¬
        """
        if not row_indices:
            return
            
        target_row = 2
        for r in row_indices:
            for c in range(1, self.last_col + 1):
                original_value = self.ws.cell(row=r, column=c).value
                ws.cell(row=target_row, column=c, value=original_value)
                
                # ì›ë³¸ ì‹œíŠ¸ì˜ ì…€ ì„œì‹ë„ ë³µì‚¬
                original_cell = self.ws.cell(row=r, column=c)
                new_cell = ws.cell(row=target_row, column=c)
                
                # í°íŠ¸ ë³µì‚¬
                if original_cell.font:
                    new_cell.font = Font(
                        color=original_cell.font.color,
                        bold=original_cell.font.bold,
                        name=original_cell.font.name,
                        size=original_cell.font.size
                    )
                
                # ë°°ê²½ìƒ‰ ë³µì‚¬
                if original_cell.fill:
                    new_cell.fill = PatternFill(
                        start_color=original_cell.fill.start_color,
                        end_color=original_cell.fill.end_color,
                        fill_type=original_cell.fill.fill_type
                    )
                
                # ì •ë ¬ ë³µì‚¬
                if original_cell.alignment:
                    new_cell.alignment = Alignment(
                        horizontal=original_cell.alignment.horizontal,
                        vertical=original_cell.alignment.vertical
                    )
            
            target_row += 1
        
        # ë°ì´í„°ê°€ 2í–‰ ì´ìƒ ìˆëŠ” ê²½ìš° ì •ë ¬ ìˆ˜í–‰
        if target_row > 3:  # í—¤ë”(1í–‰) + ë°ì´í„°(2í–‰ ì´ìƒ)
            ex = ExcelHandler(ws)
            ex.sort_by_columns([2, 3])  # Bì—´, Cì—´ ê¸°ì¤€ ì •ë ¬
        
        # Aì—´ ìˆœë²ˆ ì¬ë¶€ì—¬ "=ROW()-1"
        for r in range(2, target_row):
            ws[f"A{r}"].value = "=ROW()-1"

    def apply_automation_logic(self, ws: Worksheet) -> None:
        """ìë™í™” ë¡œì§ ì ìš© (VBA ë§¤í¬ë¡œ 21ë‹¨ê³„ê¹Œì§€ì˜ ëª¨ë“  ë¡œì§)"""
        # Excel í•¸ë“¤ëŸ¬ë¡œ ê¸°ë³¸ ì²˜ë¦¬ ì ìš©
        ex = ExcelHandler(ws)
        
        # 1. ê¸°ë³¸ ì„œì‹ ì ìš©
        ex.set_basic_format()
        
        # 2. Câ†’B ì •ë ¬
        ex.sort_by_columns([2, 3])
        
        # 3. Dì—´ ìˆ˜ì‹ ì„¤ì • (ì‚¬ìš©ì ìš”êµ¬ì‚¬í•­ì— ë§ëŠ” P, V ì²˜ë¦¬)
        self._calculate_d_column_custom(ws)
        
        # 4. ì‚¬ì´íŠ¸ë³„ ë°°ì†¡ë¹„ ì²˜ë¦¬
        ETCDeliveryFeeHandler(ws).process_delivery_fee()
        
        # 5. ì£¼ë¬¸ë²ˆí˜¸ ì²˜ë¦¬
        process_order_numbers(ws)
        
        # 6. ì „í™”ë²ˆí˜¸ ì²˜ë¦¬
        for row in range(2, ws.max_row + 1):
            for col in ('H', 'I'):
                cell_value = ws[f'{col}{row}'].value
                ws[f'{col}{row}'].value = ex.format_phone_number(cell_value)
        
        # 7. íŠ¹ìˆ˜ ì¼€ì´ìŠ¤ ì²˜ë¦¬
        special = ETCSpecialCaseHandler(ws)
        special.process_kakao_jeju()
        special.process_l_column()
        
        # 8. Fì—´ í…ìŠ¤íŠ¸ ì •ë¦¬
        for row in range(2, ws.max_row + 1):
            ws[f"F{row}"].value = ETCOrderUtils.clean_order_text(ws[f"F{row}"].value)

        # 9. ë¬¸ìì—´â†’ìˆ«ì ë³€í™˜ 
        ex.convert_numeric_strings(cols=("F","M", "W", "AA"))

        # 10. ì—´ ì •ë ¬
        ex.set_column_alignment()
        # Fì—´ ì™¼ìª½ì •ë ¬ 
        for row in range(1, ws.max_row + 1):
            ws[f"F{row}"].alignment = Alignment(horizontal='left')

        # 11. ë°°ê²½Â·í…Œë‘ë¦¬ ì œê±°, Aì—´ ìˆœë²ˆ ì„¤ì •
        self.set_row_number(ws)  # ìì²´ ì •ì˜í•œ ë©”ì„œë“œ ì‚¬ìš©
        ex.clear_fills_from_second_row()
        ex.clear_borders()

    def _calculate_d_column_custom(self, ws: Worksheet) -> None:
        """
        Dì—´ ê³„ì‚°: U + V(ìŠ¬ë˜ì‹œ í•©ì‚°)
        """
        for row in range(2, ws.max_row + 1):
            u_val = ws[f'U{row}'].value or 0
            v_val = ws[f'V{row}'].value or 0
            
            # Vì—´ ì²˜ë¦¬: "/" êµ¬ë¶„ìê°€ ìˆìœ¼ë©´ ëª¨ë“  ìˆ«ìë¥¼ í•©ì‚° (Pì—´ê³¼ ë™ì¼)
            v_num = 0
            if v_val and "/" in str(v_val):
                v_parts = str(v_val).split("/")
                for part in v_parts:
                    try:
                        v_num += float(part.strip())
                    except (ValueError, TypeError):
                        pass  # ìˆ«ìê°€ ì•„ë‹Œ ê²½ìš° ë¬´ì‹œ
            else:
                try:
                    v_num = float(v_val) if v_val else 0
                except (ValueError, TypeError):
                    v_num = 0
            
            # Dì—´ì— ê³„ì‚° ê²°ê³¼ ì„¤ì •
            calculated_d = u_val + v_num
            ws[f'D{row}'].value = calculated_d

    def _split_slash_values(self, value, expected_count: int) -> List[str]:
        """
        "/" êµ¬ë¶„ìë¡œ ë‚˜ë‰œ ê°’ë“¤ì„ ë¦¬ìŠ¤íŠ¸ë¡œ ë¶„í• 
        expected_countë§Œí¼ ê°’ì´ ì—†ìœ¼ë©´ ë¹ˆ ë¬¸ìì—´ë¡œ ì±„ì›€
        """
        if not value:
            return [""] * expected_count
        
        val_str = str(value).strip()
        if "/" in val_str:
            parts = val_str.split("/")
            # íŒŒíŠ¸ê°€ expected_countë³´ë‹¤ ì ìœ¼ë©´ ë¹ˆ ë¬¸ìì—´ë¡œ ì±„ì›€
            while len(parts) < expected_count:
                parts.append("")
            result = [part.strip() for part in parts[:expected_count]]
        else:
            # "/" êµ¬ë¶„ìê°€ ì—†ìœ¼ë©´ ì²« ë²ˆì§¸ ê°’ë§Œ ì‚¬ìš©í•˜ê³  ë‚˜ë¨¸ì§€ëŠ” ë¹ˆ ë¬¸ìì—´
            result = [val_str] + [""] * (expected_count - 1)
        
        return result

    def copy_to_new_sheet(self, 
                         wb, 
                         sheet_name: str, 
                         row_indices: List[int] = None) -> None:
        """ì§€ì •ëœ í–‰ë“¤ë¡œ ìƒˆ ì‹œíŠ¸ ìƒì„± (ë°ì´í„°ê°€ ì—†ì–´ë„ ë¹ˆ ì‹œíŠ¸ ìƒì„±)"""
        new_ws = self.create_empty_sheet(wb, sheet_name)
        if row_indices:
            self.copy_sheet_data(new_ws, row_indices)
            
        # ëª¨ë“  ì‹œíŠ¸ì— ìë™í™” ë¡œì§ ì ìš©
        self.apply_automation_logic(new_ws)

    def set_row_number(self, ws: Worksheet, start_row: int = 2) -> None:
        """
        Aì—´ ìˆœë²ˆ ìë™ ìƒì„± (=ROW()-1) - í˜„ì¬ ì‹œíŠ¸ì˜ ì‹¤ì œ í–‰ ìˆ˜ë¥¼ ê¸°ì¤€ìœ¼ë¡œ ì²˜ë¦¬
        ë¶„í• ëœ í–‰ë“¤ì„ í¬í•¨í•œ ëª¨ë“  í–‰ì— ìˆœë²ˆì„ ì •í™•íˆ ì„¤ì •
        """
        # í˜„ì¬ ì‹œíŠ¸ì˜ ì‹¤ì œ ìµœëŒ€ í–‰ ìˆ˜ë¥¼ ë™ì ìœ¼ë¡œ í™•ì¸
        end_row = ws.max_row
        
        for row in range(start_row, end_row + 1):
            ws[f'A{row}'].number_format = 'General'
            ws[f"A{row}"].value = "=ROW()-1"

    def create_automation_sheet(self, wb) -> None:
        """
        ë§¤í¬ë¡œê°€ ì ìš©ëœ ì „ì²´ ì‹œíŠ¸ë¥¼ "ìë™í™”"ë¼ëŠ” ì´ë¦„ìœ¼ë¡œ ë§¨ ì•ì— ë³µì‚¬
        """
        # ê¸°ì¡´ "ìë™í™”" ì‹œíŠ¸ê°€ ìˆìœ¼ë©´ ì‚­ì œ
        if "ìë™í™”" in wb.sheetnames:
            del wb["ìë™í™”"]
        
        # í˜„ì¬ ì›Œí¬ì‹œíŠ¸ë¥¼ ë³µì‚¬í•˜ì—¬ "ìë™í™”" ì‹œíŠ¸ ìƒì„±
        automation_ws = wb.copy_worksheet(self.ws)
        automation_ws.title = "ìë™í™”"
        
        # "ìë™í™”" ì‹œíŠ¸ë¥¼ ë§¨ ì•ìœ¼ë¡œ ì´ë™
        wb.move_sheet(automation_ws, offset=-len(wb.sheetnames) + 1)

if __name__ == "__main__":
    excel_file_path = "/Users/smith/Documents/github/OKMart/sabangnet_API/files/test-[ê¸°ë³¸ì–‘ì‹]-í•©í¬ì¥ìš©.xlsx"
    processed_file = etc_site_merge_packaging(excel_file_path)