import pandas as pd
from openpyxl import load_workbook
import os
import warnings

def reform_order_ali(file_path):
    # UserWarning 유형의 경고를 무시합니다. (openpyxl 관련 경고)
    warnings.filterwarnings("ignore", category=UserWarning)

    try:
        workbook = load_workbook(file_path, data_only=False)
        ws = workbook.active

        # '주문 ID' (A열), '주문 메모', 'Q' 열을 문자열(str)로 강제 지정하여 읽기
        df = pd.read_excel(
            file_path,
            sheet_name=ws.title,
            header=0,
            dtype={'주문 ID': str, '주문 메모': str, '상품 ID': str, '우편번호':str, '연락처':str, '모바일':str}
        )

        # 숫자로 변환 가능한 경우에만 변환하고, 그렇지 않으면 원본 값 유지하는 헬퍼 함수
        def convert_if_numeric_string(val):
            # NaN (결측치)이거나 None인 경우 그대로 반환 (openpyxl이 빈 셀로 처리)
            if pd.isna(val):
                return val
            try:
                # 숫자로 변환 시도
                return float(val)
            except (ValueError, TypeError):
                # 변환 실패 시 원본 값 (문자열 등)을 그대로 반환
                return val

        # === [1단계] H, I, L, N 열 텍스트 숫자 → 숫자 변환 및 서식 '일반' 적용 ===
        col_letters_step1 = ["H", "I", "L", "N"]
        for col_letter in col_letters_step1:
            col_idx_0based = ord(col_letter) - ord('A') # Pandas 0-based index
            col_idx_1based = col_idx_0based + 1 # Openpyxl 1-based index

            # --- 원화 기호('₩') 및 쉼표(,) 제거 ---
            df.iloc[:, col_idx_0based] = df.iloc[:, col_idx_0based].astype(str).str.replace('₩', '', regex=False).str.replace(',', '', regex=False)

            # --- 숫자로 변환 가능한 경우에만 변환 ---
            df.iloc[:, col_idx_0based] = df.iloc[:, col_idx_0based].apply(convert_if_numeric_string)

            # --- Openpyxl을 통해 해당 열의 셀 서식을 'General'로 설정 ---
            # 엑셀의 데이터 시작 행 (헤더 다음 행, 보통 2행)부터 마지막 데이터가 있는 행까지
            for row_num in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_idx_1based)
                cell.number_format = 'General'


        # === [2단계] U열 수령인주소 정리 ===
        u_col_idx_0based = ord('U') - ord('A')
        df.iloc[:, u_col_idx_0based] = df.iloc[:, u_col_idx_0based].fillna('').astype(str)
        df.iloc[:, u_col_idx_0based] = df.iloc[:, u_col_idx_0based].str.replace('대한민국、', '', regex=False)
        df.iloc[:, u_col_idx_0based] = df.iloc[:, u_col_idx_0based].str.replace('、', ' ', regex=False)

        # === [3단계] W열 구매자 국가에 "1" 입력 ===
        w_col_idx_0based = ord('W') - ord('A')
        df.iloc[:, w_col_idx_0based] = "1"

        # === [4단계] W열 텍스트 숫자 → 숫자 변환 (삼각형 제거) 및 서식 '일반' 적용 ===
        # 원화 기호가 W열에도 있을 수 있으므로 추가
        w_col_idx_1based = w_col_idx_0based + 1 # Openpyxl 1-based index

        df.iloc[:, w_col_idx_0based] = df.iloc[:, w_col_idx_0based].astype(str).str.replace('₩', '', regex=False).str.replace(',', '', regex=False)
        df.iloc[:, w_col_idx_0based] = df.iloc[:, w_col_idx_0based].apply(convert_if_numeric_string)

        # --- Openpyxl을 통해 W열의 셀 서식을 'General'로 설정 ---
        for row_num in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=w_col_idx_1based)
            cell.number_format = 'General'


        # 엑셀 워크북에 변경된 데이터 덮어쓰기 (기존 서식 유지)
        # 이 루프는 pandas DataFrame의 모든 데이터를 openpyxl 워크시트에 씁니다.
        # 기존 셀의 값만 변경하고 서식은 openpyxl 객체가 로드한 그대로 유지합니다.
        # 위에서 명시적으로 'General'로 설정된 컬럼들은 그 서식이 적용됩니다.
        for r_idx, row_data in enumerate(df.values):
            for c_idx, value in enumerate(row_data):
                cell = ws.cell(row=r_idx + 2, column=c_idx + 1)
                cell.value = value

        base, ext = os.path.splitext(file_path)
        new_file_path = f"{base}_reformed{ext}"
        workbook.save(new_file_path)
    except FileNotFoundError:
        print(f"오류: 파일을 찾을 수 없습니다: {file_path}")
    except Exception as e:
        print(f"처리 중 오류 발생: {e}")
    finally:
        return new_file_path

if __name__ == "__main__":
    excel_file = "/home/okuser/project/sabangnet_API/test_macro/o_ali_data.xlsx"
    reform_order_ali(excel_file)
