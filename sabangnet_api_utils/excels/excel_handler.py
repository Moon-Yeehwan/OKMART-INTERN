from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border
import re
import pandas as pd
from openpyxl.utils import get_column_letter
import traceback
from utils.logs.sabangnet_logger import get_logger

"""
주문관리 Excel 파일 매크로 공통 처리 메소드
- 기본 서식 설정
- 수식 처리
- 데이터 정리
- 정렬 및 레이아웃
- 특수 처리 (제주도 주소, 결제방식 등)
"""


class ExcelHandler:
    def __init__(self, ws, wb=None):
        self.ws = ws
        self.wb = wb
        self.last_row = ws.max_row

    @classmethod
    def from_file(cls, file_path, sheet_index=0):
        """
        파일 경로로 부터 엑셀 파일 로드
        예시:
            ex = ExcelHandler.from_file(file_path)
            ws = ex.ws
            wb = ex.wb
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.worksheets[sheet_index]
        return cls(ws, wb)
    
    def save_file(self, file_path):
        """
        엑셀 파일 저장
        예시:
            ex.save_file('file.xlsx')
        """
        if file_path.endswith('_매크로_완료.xlsx'):
            output_path = file_path
        else:
            output_path = file_path.replace('.xlsx', '_매크로_완료.xlsx')
        self.wb.save(output_path)
        return output_path

    def happojang_save_file(self, output_dir="files/excel/happojang", base_name=None, suffix="_매크로_완료"):
        """
        엑셀 파일 저장 (happojang 규칙 적용)
        
        Args:
            output_dir: 저장할 디렉토리 (프로젝트 루트 기준)
            base_name: 기본 파일명 (확장자 제외). None이면 현재 시트명 사용
            suffix: 파일명 접미사
            
        Returns:
            str: 저장된 파일의 전체 경로
            
        예시:
            ex.save_file()  # 기본값으로 저장
            ex.save_file(base_name="브랜디_주문데이터")  # 특정 파일명으로 저장
        """
        from pathlib import Path
        
        # 기본 파일명 결정
        if base_name is None:
            base_name = self.ws.title or "output"
            
        # 출력 디렉토리 생성
        project_root = Path(__file__).parent.parent
        output_path_dir = project_root / output_dir
        output_path_dir.mkdir(parents=True, exist_ok=True)
        
        # 최종 파일 경로
        output_path = str(output_path_dir / f"{base_name}{suffix}.xlsx")
        
        # 파일 저장
        self.wb.save(output_path)
        return output_path
    
    def set_auto_filter(self, ws=None):
        """
        A1 행 자동 필터 설정
        """
        if ws is None:
            ws = self.ws
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # 기본 서식 설정 Method
    def set_basic_format(self, ws=None, header_rgb="006100"):
        """
        폰트, 행높이, 첫 행 배경색, 줄바꿈 해제 등 기본 서식 적용
        예시:
            wb = openpyxl.load_workbook('file.xlsx')
            ws = wb.active
            set_basic_format(ws)
        """
        if ws is None:
            ws = self.ws
        font = Font(name='맑은 고딕', size=9)
        green_fill = PatternFill(start_color=header_rgb,
                                 end_color=header_rgb, fill_type="solid")
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = Alignment(wrap_text=False)
            ws.row_dimensions[row[0].row].height = 15
        for cell in ws[1]:
            cell.fill = green_fill
            cell.alignment = Alignment(horizontal='center')

    # 수식 처리 Method

    def autofill_d_column(self, ws=None, start_row=2, end_row=None, formula=None):
        """
        D열 수식 활성화 및 복사 (금액 계산)
        예시:
            autofill_d_column(ws, 2, last_row, "=U{row}+V{row}")
        - formula에 "{row}"를 포함하면 각 행 번호로 치환하여 적용
        """

        if ws is None:
            ws = self.ws
        if not end_row:
            end_row = self.last_row
        if not formula:
            formula = ws['D2'].value
        for row in range(start_row, end_row + 1):
            if ws[f'D{row}'].value is None:
                continue
            # D열 숫자 포맷 초기화
            ws[f'D{row}'].number_format = 'General'

            # 수식 적용
            if isinstance(formula, str) and '{row}' in formula:
                ws[f'D{row}'].value = formula.format(row=row)
            elif isinstance(formula, str) and '=' in formula:
                ws[f'D{row}'].value = formula.replace('2', str(row))
            else:
                ws[f'D{row}'].value = formula

    def set_row_number(self, ws, start_row=2, end_row=None):
        """
        A열 순번 자동 생성 (=ROW()-1)
        예시:
            set_row_number(ws)
        """
        if not end_row:
            end_row = self.last_row
        if ws is None:
            ws = self.ws
        for row in range(start_row, end_row + 1):
            ws[f'A{row}'].number_format = 'General'
            ws[f"A{row}"].value = "=ROW()-1"

    def convert_formula_to_value(self):
        """
        수식 → 값 변환 처리 (모든 시트)
        (openpyxl은 수식 결과값을 계산하지 않으므로, 실제 값 변환은 Excel에서 복사-값붙여넣기로 처리)
        예시:
            convert_formula_to_value(ws)
        """
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    # 실제 계산은 Excel에서, 여기선 수식 문자열만 제거
                    cell.value = cell.value

    # 데이터 정리 Method

    def clear_borders(self, ws=None):
        """
        테두리 제거 & 격자 제거
        예시:
            clear_borders(ws)
        """
        if ws is None:
            ws = self.ws
        for row in ws.iter_rows():
            ws.sheet_view.showGridLines = False
            for cell in row:
                cell.border = Border()

    def clear_fills_from_second_row(self):
        """
        배경색 제거
        예시:
            clear_fills_from_second_row(ws)
        - 두번째 행부터 모든 셀의 배경색을 제거합니다.
        """
        for row in self.ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = PatternFill(fill_type=None)

    def format_phone_number(self, val):
        """
        전화번호 포맷팅
        - 11자리 (010, 011, 016, 017, 018, 019): 010-0000-0000 형식
        - 12자리 (050, 070): 0508-0000-0000 형식  
        - 10자리 (02, 031~064): 02-0000-0000 또는 031-000-0000 형식
        예시:
            ws['H2'].value = format_phone_number(ws['H2'].value)
        """
        if not val:
            return ""
        
        val = str(val).replace('-', '').replace(' ', '').strip()
        if not val.isdigit():
            return val
        
        # 12자리: 050, 070 등
        if len(val) == 12 and val[:3] in ['050', '070']:
            return f"{val[:4]}-{val[4:8]}-{val[8:]}"
        
        # 11자리: 010, 011, 016, 017, 018, 019
        elif len(val) == 11 and val[:3] in ['010', '011', '016', '017', '018', '019']:
            return f"{val[:3]}-{val[3:7]}-{val[7:]}"
        
        # 10자리: 02(서울), 031~064(지역번호)
        elif len(val) == 10:
            if val.startswith('02'):
                return f"{val[:2]}-{val[2:6]}-{val[6:]}"
            elif val[:3] in [f'0{i}' for i in range(31, 65)]:
                return f"{val[:3]}-{val[3:6]}-{val[6:]}"
        
        # 9자리: 02 + 7자리 (서울 구번호)
        elif len(val) == 9 and val.startswith('02'):
            return f"{val[:2]}-{val[2:5]}-{val[5:]}"
        
        return val

    def clean_model_name(self, val):
        """
        모델명에서 ' 1개' 텍스트 제거
        예시:
            ws['F2'].value = clean_model_name(ws['F2'].value)
        """
        return str(val).replace(' 1개', '') if val else val

    def sum_prow_with_slash(self):
        """
        P열 "/" 금액 합산 (음수 지원)
        예시: "216/-56" → 160
            sum_prow_with_slash(ws)
        """
        last_row = self.ws.max_row
        for r in range(2, last_row + 1):
            p_cell = self.ws[f"P{r}"]
            p_raw = p_cell.value
            
            # 이미 숫자인 경우 그대로 유지
            if isinstance(p_raw, (int, float)):
                continue
                
            p_str = str(p_raw or "").strip()
            if "/" in p_str:
                nums = []
                for n in p_str.split("/"):
                    n = n.strip()
                    if n:  # 빈 문자열이 아닌 경우
                        try:
                            # 숫자 변환 시도 (음수 포함)
                            nums.append(float(n))
                        except ValueError:
                            # 숫자가 아닌 경우 to_num 메서드로 처리
                            converted = self.to_num(n)
                            if converted != 0 or n == '0':  # 0은 유효한 값으로 처리
                                nums.append(converted)
                p_cell.value = sum(nums) if nums else p_raw  # 변환 실패시 원본 유지
            elif p_str and p_str != "0":  # 빈 값이 아니고 "0"이 아닌 경우만 변환
                converted = self.to_num(p_str)
                if converted != 0:  # 변환 결과가 0이 아닌 경우만 적용
                    p_cell.value = converted

    def to_num(self, val) -> int:
        """
        '12,345원' → 12345.0 (실패 시 0)
        예시:
            num = to_num("12,345원")
        """
        try:
            return int(re.sub(r"[^\d.-]", "", str(val))) if str(val).strip() else 0.0
        except ValueError:
            return 0.0

    def convert_numeric_strings(self, ws=None, start_row: int = 2, end_row: int | None = None, cols: tuple[str, ...] | None = None) -> None:
        """
        워크시트의 문자열 숫자를 숫자 타입(int/float)으로 변환합니다.

        Args:
            start_row: 변환을 시작할 데이터 행 (헤더가 1행이라고 가정).
            end_row: 변환을 끝낼 행. None이면 마지막 행까지 처리.
            cols: 변환할 열 머리글(예: ("E","M","Q","W")). None이면 모든 열을 대상.

        사용 예시:
            # 1) 모든 열 대상
            ex.convert_numeric_strings()

            # 2) 특정 열만
            ex.convert_numeric_strings(cols=("E", "M", "Q", "W"))
        """
        if ws is None:
            ws = self.ws
        if end_row is None:
            end_row = self.ws.max_row

        # 변환 대상 열 결정
        if cols:
            target_cols = cols
        else:
            # 1행 헤더를 기준으로 모든 실제 열 레터를 수집
            target_cols = tuple(
                cell.column_letter for cell in ws[1] if cell.value is not None)

        for row in range(start_row, end_row + 1):
            for col in target_cols:
                cell = ws[f"{col}{row}"]
                if cell.value is None:
                    continue
                if isinstance(cell.value, str):
                    raw = cell.value.strip()
                    # 숫자(0-9), 쉼표, 마침표 외 다른 문자가 섞여 있으면 변환하지 않음
                    if re.fullmatch(r"[0-9,\.]+", raw):
                        num_val = self.to_num(raw)
                        # 0 도 유효 숫자로 인정
                        if raw not in {"", ".", ","}:
                            cell.value = num_val
                            cell.number_format = "0"

    # 정렬 및 레이아웃 Method

    def set_column_alignment(self, ws=None):
        """
        A,B(가운데), D,E,G(오른쪽), 첫 행 가운데 정렬
        예시:
            set_column_alignment(ws)
        """
        center = Alignment(horizontal='center')
        right = Alignment(horizontal='right')

        align_map = {
            'center': {'A', 'B'},
            'right': {'D', 'E', 'G'}
        }
        if ws is None:
            ws = self.ws
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                col_letter = cell.column_letter
                if col_letter in align_map['center']:
                    cell.alignment = center
                elif col_letter in align_map['right']:
                    cell.alignment = right


    def sort_dataframe_by_c_b(self, df, c_col='C', b_col='B'):
        """
        DataFrame을 C열 → B열 순서로 오름차순 정렬
        예시:
            df = sort_dataframe_by_c_b(df)
        """
        if c_col in df.columns and b_col in df.columns:
            print(c_col, b_col)
            return df.sort_values(by=[c_col, b_col]).reset_index(drop=True)
        return df
    
    def sort_by_columns(self, key_columns: List[int], start_row: int = 2) -> None:
        """
        지정된 열들을 기준으로 워크시트 데이터 정렬
        
        :param key_columns: 정렬 기준 열 번호 리스트 (1-based indexing)
                        예: [2, 3]은 B열, C열 순서로 정렬
        :param start_row: 정렬 시작 행 번호 (기본값: 2, 첫 행은 헤더)
        
        예시:
            # B열, C열 순서로 2단계 정렬
            ex = ExcelHandler.from_file("example.xlsx")
            ex.sort_by_columns([2, 3])
            
            # D열 기준 단일 정렬, 3행부터
            ex.sort_by_columns([4], start_row=3)
            
            # 여러 열 조합 정렬 (A → C → B)
            ex.sort_by_columns([1, 3, 2])
        
        주의:
        - 열 번호는 1부터 시작 (A열=1, B열=2, ...)
        - 정렬은 문자열 비교 기준 ('123' > '1000')
        - 정렬 후 자동으로 행 번호 재설정되지 않음
        필요시 set_row_number() 별도 호출
        """
        rows = [
            [self.ws.cell(row=r, column=c).value 
            for c in range(1, self.ws.max_column + 1)]
            for r in range(start_row, self.last_row + 1)
        ]
        
        # 정렬 키 함수: 각 열을 문자열로 변환하여 비교
        rows.sort(key=lambda x: tuple(str(x[i-1]) for i in key_columns))
        
        # 기존 데이터 삭제 후 정렬된 데이터 재기록
        self.ws.delete_rows(start_row, self.last_row - start_row + 1)
        for ridx, row in enumerate(rows, start=start_row):
            for cidx, val in enumerate(row, start=1):
                self.ws.cell(row=ridx, column=cidx, value=val)
        
        # last_row 업데이트
        self.last_row = self.ws.max_row

    # 특수 처리 Method

    def process_jeju_address(self, row,ws=None, f_col='F', j_col='J'):
        """
        제주도 주소: '[3000원 연락해야함]' 추가, 연한 파란색 배경 및 빨간 글씨 적용
        예시:
            process_jeju_address(ws, row=5)
        """
        if ws is None:
            ws = self.ws
        red_font = Font(color="FF0000", bold=True)
        # RGB(204,255,255) → hex: "CCFFFF"
        light_blue_fill = PatternFill(
            start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
        # F열 안내문 추가
        f_val = ws[f'{f_col}{row}'].value
        if f_val and "[3000원 연락해야함]" not in str(f_val):
            ws[f'{f_col}{row}'].value = str(f_val) + " [3000원 연락해야함]"
        # J열 빨간 글씨
        ws[f'{j_col}{row}'].font = red_font
        # F열 연한 파란색 배경
        ws[f'{f_col}{row}'].fill = light_blue_fill

    def process_l_column(self, row, l_col='L'):
        """
        L열 결제방식: '신용' 삭제, '착불' 빨간 글씨
        예시:
            process_l_column(ws, row=7)
        """
        red_font = Font(color="FF0000", bold=True)
        l_val = self.ws[f'{l_col}{row}'].value
        if l_val == "신용":
            self.ws[f'{l_col}{row}'].value = ""
        elif l_val == "착불":
            self.ws[f'{l_col}{row}'].font = red_font

    def highlight_column(self, col: str, light_color: PatternFill, ws=None, start_row: int = 2, last_row: int = None):
        """
        특정 열 하이라이트 처리
        예시:
            - F열 모르겠는 셀 색칠음영 (하늘색)
            - highlight_column(col='F', light_color=light_blue_fill, start_row=2, last_row=last_row)
        """

        def _should_highlight(txt: str) -> bool:
            """
            셀 값이 다음 조건 중 하나라도 만족하면 True:
            - 빈 문자열
            - 'none' (대소문자 무관)
            - 모든 문자가 '#'
            - 순수 숫자
            - '숫자개' 패턴 (예: '3개')
            """
            if not txt or txt.strip() == "":
                return True
            txt = txt.strip()
            if txt.lower() == "none":
                return True
            if all(c == '#' for c in txt):
                return True
            if txt.isdigit():
                return True
            if txt.endswith("개") and txt[:-1].isdigit():
                return True
            return False

        if ws is None:
            ws = self.ws
        if not last_row:
            last_row = self.last_row
        for row in range(start_row, last_row + 1):
            cell_value = ws[f'{col}{row}'].value
            txt = str(cell_value).strip() if cell_value else ""

            if cell_value is not None and _should_highlight(txt):
                ws[f"{col}{row}"].fill = light_color

    def set_header_style(self, ws):
        """
        지정한 워크시트의 헤더 행에 배경색, 폰트, 정렬을 일괄 적용
        Args:
            ws: openpyxl worksheet
            headers: 헤더 리스트
            fill: 헤더 배경색(hex)
            font: 폰트
            alignment: 정렬 
        """
        white_font = Font(name='맑은 고딕', size=9, color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal='center')
        green_fill = PatternFill(start_color="008000",
                                 end_color="008000", fill_type="solid")
        for cell in ws[1]:
            cell.fill = green_fill
            cell.font = white_font
            cell.alignment = center_alignment
            cell.border = Border()
            ws.row_dimensions[1].height = 15

    def convert_to_number(self, cell_value):
        """
        문자열을 숫자로 변환
        예시:
            convert_to_number(ws['M2'].value)
        """
        return float(cell_value) if '.' in str(cell_value) else int(float(cell_value))

    def to_dataframe(self, ws=None, start_row=2, start_col=1, end_row=None, end_col=None):
        """
        지정된 워크시트의 데이터를 DataFrame으로 변환
        args:
            ws: 워크시트
            start_row: 시작 행
            start_col: 시작 열
            end_row: 끝 행
            end_col: 끝 열
        """
        ws = ws or self.ws
        end_row = end_row or ws.max_row
        end_col = end_col or ws.max_column

        # 헤더 추출
        headers = []

        for col in range(start_col, end_col + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(header if header else f"Col{col}")

        # 데이터 추출
        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                row_data.append(ws.cell(row=row, column=col).value)
            data.append(row_data)

        return pd.DataFrame(data, columns=headers)

    def create_split_sheets(self, headers: list, sheet_names: list):
        """
        지정한 이름의 시트를 생성하고, 열 너비/행 높이만 원본 시트(self.ws)에서 복사합니다.
        헤더 복사 및 스타일 적용은 제외합니다.

        Args:
            headers (list): 헤더 리스트 
            sheet_names (list): 생성할 시트명 리스트 ["OK,CL,BB", "IY"]

        Returns:
            dict: {시트명: 워크시트 객체}
        """
        ws_map = {}
        for sheet_name in sheet_names:
            # 기존 시트 삭제
            if sheet_name in self.wb.sheetnames:
                del self.wb[sheet_name]
            # 새 시트 생성
            ws = self.wb.create_sheet(title=sheet_name)
            # 열 너비 복사
            for col in range(1, len(headers) + 1):
                col_letter = get_column_letter(col)
                src_width = self.ws.column_dimensions[col_letter].width
                ws.column_dimensions[col_letter].width = src_width
            # 행 높이 복사 (헤더 행만)
            ws.row_dimensions[1].height = 15
            ws_map[sheet_name] = ws

        return ws_map

    def split_sheets_by_site(self, df, ws_map, site_mapping):
        """
        공통 시트 분리 메서드
        
        Args:
            rules (dict): 간단한 규칙 딕셔너리
                        예: {"OK": ["오케이마트"], "IY": ["아이예스"], "OK,CL,BB": ["오케이마트", "클로버프", "베이지베이글"]}
        """
        # 각 시트별 행 인덱스 초기화
        site_rows = {sheet: 2 for sheet in site_mapping.keys()}
        font = Font(name='맑은 고딕', size=9)
        
        for row_data in df.itertuples(index=False):
            # 계정명 추출
            site_value = str(getattr(row_data, '사이트')) if pd.notna(getattr(row_data, '사이트')) else ""
            account_name = ""
            
            if "]" in site_value and site_value.startswith("["):
                try:
                    account_name = site_value[1:site_value.index("]")]
                except:
                    account_name = ""
            
            # 매칭되는 시트 찾기
            for sheet, filters in site_mapping.items():
                if account_name in filters and sheet in ws_map:
                    target_sheet = ws_map[sheet]
                    current_row = site_rows[sheet]
                    
                    # 데이터 복사
                    for col_idx, value in enumerate(row_data, 1):
                        cell = target_sheet.cell(row=current_row, column=col_idx, value=value)
                        cell.font = font
                    
                    target_sheet.row_dimensions[current_row].height = 15
                    site_rows[sheet] += 1
                    break

    @staticmethod
    def from_upload_file_to_dataframe(upload_file, sheet_index=0, **to_df_kwargs):
        """
        업로드 파일(UploadFile 등 file-like object)을 임시 파일로 저장하고, DataFrame으로 읽고, 임시 파일 삭제 후 DataFrame 반환
        Args:
            upload_file: FastAPI UploadFile 등 file-like object
            sheet_index: 읽을 시트 인덱스(기본 0)
            **to_df_kwargs: to_dataframe에 전달할 추가 인자
        Returns:
            pd.DataFrame
        """
        import tempfile
        import os
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp.write(upload_file.file.read())
                tmp_path = tmp.name
            ex = ExcelHandler.from_file(tmp_path, sheet_index=sheet_index)
            df = ex.to_dataframe(**to_df_kwargs)
        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
        return df

    def calculate_d_column_values(self, ws=None, start_row=2, end_row=None, first_col=None, second_col=None, third_col=None):
        """
        D열에 지정된 열들의 실제 계산된 값을 입력 (수식이 아닌 값)
        - 2개 열: first_col + second_col
        - 3개 열: first_col + second_col + third_col
        
        예시:
            # 3개 열 합산 (O+P+V)
            calculate_d_column_values(ws, first_col='O', second_col='P', third_col='V')
            
            # 2개 열 합산 (M+N)
            calculate_d_column_values(ws, first_col='M', second_col='N')
            
            # 다른 열 조합 (U+W+X)
            calculate_d_column_values(ws, first_col='U', second_col='W', third_col='X')
            
            # 행 범위 지정
            calculate_d_column_values(ws, start_row=3, end_row=100, first_col='A', second_col='B')
        """
        if ws is None:
            ws = self.ws
        if not end_row:
            end_row = self.last_row
            
        # 최소 2개 열은 필요
        if first_col is None or second_col is None:
            raise ValueError("first_col과 second_col은 필수 파라미터입니다.")
        
        for row in range(start_row, end_row + 1):
            # 각 열의 원본 값 확인
            first_raw = ws[f'{first_col}{row}'].value
            second_raw = ws[f'{second_col}{row}'].value
            third_raw = ws[f'{third_col}{row}'].value if third_col else None
            
            # 각 열의 값을 가져와서 숫자로 처리
            first_val = first_raw if isinstance(first_raw, (int, float)) else 0
            second_val = second_raw if isinstance(second_raw, (int, float)) else 0
            third_val = third_raw if isinstance(third_raw, (int, float)) else 0
            
            # 2개 열 또는 3개 열 합산
            if third_col is None:
                # 2개 열 합산
                total = first_val + second_val
            else:
                # 3개 열 합산
                total = first_val + second_val + third_val
            
            # 합계를 D열에 입력
            ws[f'D{row}'].value = total
            ws[f'D{row}'].number_format = 'General'

    @staticmethod
    def file_path_to_dataframe(file_path, sheet_index=0, **to_df_kwargs):
        """
        임시파일을 DataFrame으로 읽고, 임시 파일 삭제 후 DataFrame 반환
        Args:
            file_path: 업로드 된 파일 경로
            sheet_index: 읽을 시트 인덱스(기본 0)
            **to_df_kwargs: to_dataframe에 전달할 추가 인자
        Returns:
            pd.DataFrame
        """
        import os
        from minio_handler import delete_temp_file
        try:
            ex = ExcelHandler.from_file(file_path, sheet_index=sheet_index)
            df = ex.to_dataframe(**to_df_kwargs)
        finally:
            delete_temp_file(file_path)
        return df

    def preprocess_and_update_ws(self, ws, sort_columns: list[int]):
        """
        1. 헤더/데이터 추출
        2. 데이터 정렬
        3. 원본 시트 "자동화"로 이름 변경 및 정렬된 데이터로 업데이트
        return: (headers, sorted_data)
        """
        # 1. 헤더와 데이터 추출
        headers, data = self._extract_headers_and_data(ws)

        # 2. 데이터 정렬
        data = self._sort_data(data, sort_columns)

        # 3. 원본 시트를 "자동화"로 이름 변경하고 정렬된 데이터로 업데이트
        ws.title = "자동화"
        self._update_worksheet_data(ws, data)

        return headers, data

    def split_and_write_ws_by_site(
        self,
        wb,
        headers: list,
        data: list[list],
        sheets_name: list[str],
        site_to_sheet: dict,
        site_col_idx: int = 2,
    ):
        """
        wb: 워크북
        headers: 헤더
        data: 데이터
        sheets_name: 생성할 시트명 리스트
        site_to_sheet: {사이트명: 시트명}
        site_col_idx: 사이트 컬럼 인덱스 (1-based) 기본값 2
        """

        # 4. 시트들 생성
        filtered_sheets = [name for name in sheets_name if name != "자동화"]
        ws_map = self._create_sheets(
            wb.worksheets[0], headers, filtered_sheets)

        # 5. 각 필터링된 시트에 데이터 삽입
        self._write_data_to_sheets(data, ws_map, site_to_sheet, site_col_idx)

        # 6. 컬럼 너비 복사
        self._copy_column_widths(wb)

    def _extract_headers_and_data(self, ws) -> tuple[list, list[list]]:
        """
        워크시트에서 헤더와 데이터 추출
        args:
            ws: 워크시트
        return:
            headers: 헤더
            data: 데이터
        """
        # 헤더 추출
        headers = [ws.cell(row=1, column=c).value for c in range(
            1, ws.max_column + 1)]

        # 데이터 추출
        data = [
            [ws.cell(row=r, column=c).value for c in range(
                1, ws.max_column + 1)]
            for r in range(2, ws.max_row + 1)
        ]

        return headers, data
    
    def _sort_data(self, data: list[list], sort_columns: list[int]) -> list[list]:
        """
        데이터 정렬 (컬럼 인덱스) 2025-07-17 srot_columns에 음수 값이 입력되면 역순 정렬되도록 수정
        args:
            data: 데이터
            sort_columns: 정렬 기준 컬럼 인덱스
        return:
            sorted_data: 정렬된 데이터
        """
        logger = get_logger(__name__)
        
        def dynamic_key(item):
            key_tuple_elements = []
            try:
                for col_idx in sort_columns:
                    value = item[abs(col_idx)-1]
                    # None 처리: 숫자면 inf/-inf, 문자열이면 "", 기타는 ""
                    if value is None:
                        # 오름차순: None은 맨 뒤, 내림차순: None은 맨 앞
                        if col_idx > 0:
                            value = float('inf') if isinstance(item[abs(col_idx)-1], (int, float)) else ""
                        else:
                            value = float('-inf') if isinstance(item[abs(col_idx)-1], (int, float)) else ""
                    if col_idx > 0:
                        key_tuple_elements.append(value)
                    elif col_idx < 0:
                        if isinstance(value, (int, float)):
                            key_tuple_elements.append(-value)
                        elif isinstance(value, str):
                            key_tuple_elements.append(ReverseComparableString(value))
                        else:
                            key_tuple_elements.append(value)
            except Exception as exc:
                logger.error(
                    f"_sort_data dynamic_key() 에러: "
                    f"item={item}, sort_columns={sort_columns}, "
                    f"key_tuple_elements={key_tuple_elements}"
                )
                raise
            return tuple(key_tuple_elements)
        
        try:
            return sorted(data, key=dynamic_key)
        except Exception as exc:
            logger.error(
                f"_sort_data sorted() 에러: "
                f"data_length={len(data)}, sort_columns={sort_columns}, "
                f"data_sample={data[:3] if data else 'empty'}\n{traceback.format_exc()}"
            )
            raise

    def _update_worksheet_data(self, ws, data: list[list]):
        """
        워크시트에 정렬된 데이터 업데이트
        args:
            ws: 워크시트
            data: 데이터
        """
        font = Font(name='맑은 고딕', size=9)
        empty_fill = PatternFill()

        # 정렬된 데이터 다시 삽입 및 기본 스타일 제거
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx,
                               value=value if value or value == 0 else "")
                cell.fill = empty_fill
                cell.font = font
                cell.alignment = Alignment(wrap_text=False)
                cell.border = Border()
            ws.row_dimensions[row_idx].height = 15
            ws.sheet_view.showGridLines = True

    def _create_sheets(self, ws, headers: list, filtered_sheets: list[str]) -> dict:
        """
        시트들 생성
        args:
            ws: 워크시트
            headers: 헤더
            filtered_sheets: 생성할 시트명 리스트
        """
        ws_map = {}

        for sheet_name in filtered_sheets:
            # 기존 시트 삭제 (존재하는 경우)
            if sheet_name in ws.parent.sheetnames:
                del ws.parent[sheet_name]

            new_ws = ws.parent.create_sheet(title=sheet_name)

            # 헤더 추가
            for col, header in enumerate(headers, start=1):
                new_ws.cell(row=1, column=col, value=header)

            ws_map[sheet_name] = new_ws

        return ws_map

    def _write_data_to_sheets(self, data: list[list], ws_map: dict, site_to_sheet: dict, site_col_idx: int = 2):
        """
        데이터를 스트리밍 방식으로 각 시트에 삽입
        args:
            data: 데이터
            ws_map: 시트 매핑
            site_to_sheet: 사이트 매핑
        """
        account_pattern = re.compile(r'^\[([^\]]+)\]')
        empty_fill = PatternFill()
        font = Font(name='맑은 고딕', size=9)

        for row in data:
            # 사이트 정보 추출
            site_value = str(
                row[site_col_idx - 1]) if len(row) > site_col_idx - 1 and row[site_col_idx - 1] else ""

            # 계정명 추출
            match = account_pattern.match(site_value)
            if match:
                account_name = match.group(1)
                target_sheet_name = site_to_sheet.get(account_name)

                # 해당 시트에 즉시 데이터 삽입
                if target_sheet_name and target_sheet_name in ws_map:
                    target_ws = ws_map[target_sheet_name]

                    # 다음 빈 행에 데이터 삽입 (max_row + 1)
                    next_row = target_ws.max_row + 1

                    # 행 데이터 삽입
                    for col_idx, value in enumerate(row, start=1):
                        cell = target_ws.cell(
                            row=next_row, column=col_idx, value=value)
                        cell.fill = empty_fill
                        cell.font = font
                        cell.alignment = Alignment(wrap_text=False)
                    # 행 높이 설정
                    target_ws.row_dimensions[next_row].height = 15

    def _copy_column_widths(self, wb):
        """
        컬럼 너비 복사
        args:
            wb: 워크북
        추가 : 너비 설정 확인 필요
        """
        from openpyxl.utils import get_column_letter

        for target_ws in wb.worksheets[1:]:
            source_ws = wb.worksheets[0]
            for col_num in range(1, source_ws.max_column + 1):
                col_letter = get_column_letter(col_num)
                # 소스 워크시트에서 컬럼 너비 가져오기
                if col_letter in source_ws.column_dimensions:
                    src_width = source_ws.column_dimensions[col_letter].width
                    target_ws.column_dimensions[col_letter].width = src_width

    def create_vlookup_dict(self, wb):
        vlookup_dict = {}
        for sheet in wb:
            if sheet.title == "Sheet1":
                for row in range(2, sheet.max_row + 1):
                   vlookup_dict[str(sheet.cell(row=row, column=1).value)] = str(sheet.cell(row=row, column=2).value)
                del wb[sheet.title]
        return vlookup_dict

class ReverseComparableString:
    def __init__(self, s):
        self.s = s

    # '작다' (<) 연산을 뒤집어 '크다' (>)로 동작하게 함
    def __lt__(self, other):
        return self.s > other.s

    # '같다' (==) 연산은 그대로
    def __eq__(self, other):
        return self.s == other.s