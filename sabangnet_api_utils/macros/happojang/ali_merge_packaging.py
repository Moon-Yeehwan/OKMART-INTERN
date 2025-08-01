"""ì•Œë¦¬ìµìŠ¤í”„ë ˆìŠ¤ í•©í¬ì¥ ìë™í™” ëª¨ë“ˆ"""

from __future__ import annotations
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from utils.excels.excel_handler import ExcelHandler


# ì„¤ì • ìƒìˆ˜
MALL_NAME = "ì•Œë¦¬ìµìŠ¤í”„ë ˆìŠ¤"
RED_FONT = Font(color="FF0000", bold=True)


class ALIProductUtils:
    # ì •ê·œì‹ íŒ¨í„´
    STAR_QTY_RE = re.compile(r"\* ?(\d+)")
    MULTI_SEP_RE = re.compile(r"[\/;]")
    DUP_QTY_RE = re.compile(r"\d+ê°œ")
    PHONE_RE = re.compile(r"\D")
    JEJU_RE = re.compile(r"ì œì£¼|ì„œê·€í¬")
    
    @staticmethod
    def clean_product_text(txt: str | None) -> str:
        """
        ğŸ”„ ExcelHandler í›„ë³´
        ìƒí’ˆëª… ë¬¸ìì—´ ì •ë¦¬
        - '/' ';' â†’ ' + '
        - '*n' â†’ 'nê°œ' (n>1)
        - ' 1ê°œ' ì œê±°
        """
        if not txt:
            return ""
            
        txt = ALIProductUtils.MULTI_SEP_RE.sub(" + ", str(txt))
        
        def qty_replace(m: re.Match) -> str:
            n = m.group(1).strip()
            return "" if n == "1" else f" {n}ê°œ"
            
        txt = ALIProductUtils.STAR_QTY_RE.sub(qty_replace, txt)
        return txt.replace(" 1ê°œ", "").strip()

    @staticmethod
    def format_phone(val: str | None) -> str:
        """ì „í™”ë²ˆí˜¸ í¬ë§· (01012345678 â†’ 010-1234-5678)"""
        if not val:
            return ""
        digits = ALIProductUtils.PHONE_RE.sub("", str(val))
        if digits.startswith("10"):
            digits = "0" + digits
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}" if len(digits) == 11 else digits

    @staticmethod
    def check_multiple_quantities(txt: str) -> bool:
        """Fì—´ ë‹¤ì¤‘ìˆ˜ëŸ‰ ì²´í¬ ('ê°œ' 2íšŒ ì´ìƒ ë“±ì¥)"""
        parts = [p.strip() for p in str(txt or "").split("+")]
        return sum(1 for p in parts if ALIProductUtils.DUP_QTY_RE.search(p)) >= 2

    @staticmethod
    def is_jeju_address(addr: str) -> bool:
        """ì£¼ì†Œê°€ ì œì£¼ë„ì¸ì§€ í™•ì¸"""
        return bool(ALIProductUtils.JEJU_RE.search(str(addr or "")))

    @staticmethod
    def build_lookup_map(ws_lookup: Worksheet) -> Dict[str, str]:
        """Sheet1ì˜ A:Bë¥¼ ë”•ì…”ë„ˆë¦¬ë¡œ ë³€í™˜"""
        return {
            str(r[0]): r[1]
            for r in ws_lookup.iter_rows(min_row=2, max_col=2, values_only=True)
            if r[0] is not None
        }


FONT_MALGUN = Font(name="ë§‘ì€ ê³ ë”•", size=9)
HDR_FILL = PatternFill(start_color="006100",
                       end_color="006100", fill_type="solid")
BLUE_FILL = PatternFill(start_color="CCE8FF",
                        end_color="CCE8FF", fill_type="solid")
JEJU_FILL = PatternFill(start_color="DDEBF7",
                        end_color="DDEBF7", fill_type="solid")
NO_BORDER = Border()


def to_num(val) -> float:
    """'12,345ì›' â†’ 12345.0 (ì‹¤íŒ¨ ì‹œ 0)."""
    try:
        return float(re.sub(r"[^\d.-]", "", str(val))) if str(val).strip() else 0.0
    except ValueError:
        return 0.0


def calculate_d_column_with_numbers(ws: Worksheet) -> None:
    """
    Dì—´ì— Uì—´ê³¼ Vì—´ì˜ ì‹¤ì œ ìˆ«ì ê°’ì„ ë”í•œ ê²°ê³¼ë¥¼ ì €ì¥
    ìˆ˜ì‹ì´ ì•„ë‹Œ ì‹¤ì œ ê³„ì‚°ëœ ìˆ«ì ë°ì´í„°ë¥¼ ì…ë ¥
    """
    for row in range(2, ws.max_row + 1):
        # Uì—´ ê°’ ê°€ì ¸ì˜¤ê¸°
        u_value = ws[f"U{row}"].value
        v_value = ws[f"V{row}"].value
        
        # ìˆ«ìë¡œ ë³€í™˜ (Noneì´ë‚˜ ë¹ˆ ê°’ì€ 0ìœ¼ë¡œ ì²˜ë¦¬)
        try:
            u_num = float(u_value) if u_value is not None else 0.0
        except (ValueError, TypeError):
            u_num = 0.0
            
        try:
            v_num = float(v_value) if v_value is not None else 0.0
        except (ValueError, TypeError):
            v_num = 0.0
        
        # Dì—´ì— ì‹¤ì œ ê³„ì‚°ëœ ìˆ«ì ì €ì¥
        calculated_result = u_num + v_num
        ws[f"D{row}"].value = calculated_result


def copy_product_info(ws: Worksheet) -> None:
    """Zì—´ ìƒí’ˆì •ë³´ë¥¼ Fì—´ë¡œ ë³µì‚¬í•˜ê³  ì •ë¦¬"""
    for row in range(2, ws.max_row + 1):
        z_val = ws[f"Z{row}"].value
        ws[f"F{row}"].value = ALIProductUtils.clean_product_text(z_val)
        if ALIProductUtils.check_multiple_quantities(ws[f"F{row}"].value):
            ws[f"F{row}"].fill = BLUE_FILL


def process_phones(ws: Worksheet) -> None:
    """ì „í™”ë²ˆí˜¸ ì²˜ë¦¬ (Iì—´ í¬ë§· + Hì—´ ë³µì‚¬)"""
    for row in range(2, ws.max_row + 1):
        phone = ALIProductUtils.format_phone(ws[f"I{row}"].value)
        ws[f"I{row}"].value = phone
        ws[f"H{row}"].value = phone
    ws.column_dimensions["H"].width = ws.column_dimensions["I"].width


def process_jeju_orders(ex: ExcelHandler) -> None:
    """ì œì£¼ë„ ì£¼ë¬¸ ì²˜ë¦¬"""
    ws = ex.ws
    for row in range(2, ws.max_row + 1):
        if ALIProductUtils.is_jeju_address(ws[f"J{row}"].value):
            ws[f"J{row}"].font = RED_FONT
            if "[3000ì› ì—°ë½í•´ì•¼í•¨]" not in str(ws[f"F{row}"].value):
                ws[f"F{row}"].value = f"{ws[f'F{row}'].value} [3000ì› ì—°ë½í•´ì•¼í•¨]"
            ws[f"F{row}"].fill = BLUE_FILL


def ali_merge_packaging(input_path: str) -> str:
    """ì•Œë¦¬ìµìŠ¤í”„ë ˆìŠ¤ ì£¼ë¬¸ í•©í¬ì¥ ìë™í™” ì²˜ë¦¬"""
    # Excel íŒŒì¼ ë¡œë“œ
    ex = ExcelHandler.from_file(input_path)
    ws = ex.ws

    # 1. ê¸°ë³¸ ì„œì‹ ì ìš©
    ex.set_basic_format()
    
    # 2. Pì—´ ìŠ¬ë˜ì‹œ(/) ê¸ˆì•¡ í•©ì‚° 
    # TODO: ì´ ë¶€ë¶„ í™•ì¸ í•„ìš”
    ex.sum_prow_with_slash()
    
    # 3. Câ†’B ì •ë ¬
    ex.sort_by_columns([2, 3])  # Cì—´=3, Bì—´=2
    
    # 4. ìƒí’ˆì •ë³´ ë³µì‚¬ ë° ì •ë¦¬ (Z â†’ F)
    copy_product_info(ws)
    
    # 5. ë°°ê²½ ì œê±°
    ex.clear_fills_from_second_row()
    
    # 7-8. ì „í™”ë²ˆí˜¸ ì²˜ë¦¬
    process_phones(ws)
    
    # 9-10. Eì—´ LEFT(16) ì²˜ë¦¬
    ws.insert_cols(6)
    ws["F1"].value = "TrimE"
    for r in range(2, ws.max_row + 1):
        ws[f"F{r}"].value = str(ws[f"E{r}"].value)[:16]
    for r in range(2, ws.max_row + 1):
        ws[f"E{r}"].value = ws[f"F{r}"].value
    ws.delete_cols(6)
    
    # 12. Dì—´ ê³„ì‚° (Uì—´ + Vì—´ì˜ ì‹¤ì œ ìˆ«ì ê°’)
    calculate_d_column_with_numbers(ws)
    
    # 13. Aì—´ ìˆœë²ˆ ì„¤ì •
    ex.set_row_number(ws)
    
    # 14-17. ì—´ ì •ë ¬/ì„œì‹
    ex.set_column_alignment()
    ex.clear_borders()
    ws.column_dimensions["E"].width = 20
    
    # 19. ì œì£¼ë„ ì£¼ë¬¸ ì²˜ë¦¬
    process_jeju_orders(ex)
    
    # E, M, P, Q, W ì—´ Stringìˆ«ì to ìˆ«ì ë³€í™˜
    ex.convert_numeric_strings(cols=("E", "M", "P", "Q", "W"))
    
    # 21. Sì—´ VLOOKUP ì²˜ë¦¬ (Sheet1ì´ ìˆëŠ” ê²½ìš°)
    if "Sheet1" in ex.wb.sheetnames:
        lookup_map = ALIProductUtils.build_lookup_map(ex.wb["Sheet1"])
        for row in range(2, ws.max_row + 1):
            m_val = str(ws[f"M{row}"].value)
            ws[f"S{row}"].value = lookup_map.get(m_val, "S")
    
    # 22. ì‹œíŠ¸ ë¶„ë¦¬ (OK, IY)
    splitter = ALISheetSplitter(ws)
    rows_by_sheet = splitter.get_rows_by_sheet()
    
    for sheet_name, row_indices in rows_by_sheet.items():
        if row_indices:  # í•´ë‹¹ ì‚¬ì´íŠ¸ì˜ ë°ì´í„°ê°€ ìˆëŠ” ê²½ìš°ë§Œ
            splitter.copy_to_new_sheet(ex.wb, sheet_name, row_indices)

    # # 24. ì‹œíŠ¸ ìˆœì„œ ì •ë¦¬
    # desired = ["ìë™í™”", "OK", "IY", "Sheet1"]
    # for name in reversed(desired):
    #     if name in ex.wb.sheetnames:
    #         ex.wb._sheets.insert(0, ex.wb._sheets.pop(ex.wb.sheetnames.index(name)))
            
    # ì €ì¥
    base_name = Path(input_path).stem  # í™•ì¥ì ì œê±°í•œ íŒŒì¼ëª…
    output_path = ex.happojang_save_file(base_name=base_name)
    
    print(f"â—¼ï¸ [{MALL_NAME}] í•©í¬ì¥ ìë™í™” ì™„ë£Œ!")
    
    return output_path


class ALISheetSplitter:
    """ì‹œíŠ¸ ë¶„ë¦¬ ì²˜ë¦¬ í´ë˜ìŠ¤"""
    
    def __init__(self, ws: Worksheet):
        self.ws = ws
        self.last_row = ws.max_row
        self.last_col = ws.max_column
        
        # ì—´ ë„ˆë¹„ ì €ì¥
        self.col_widths = [
            ws.column_dimensions[get_column_letter(c)].width
            for c in range(1, self.last_col + 1)
        ]

    def get_rows_by_sheet(self) -> Dict[str, List[int]]:
        """ì‚¬ì´íŠ¸ë³„ í–‰ ë²ˆí˜¸ ë§¤í•‘ ìƒì„±"""
        site_rows = defaultdict(list)
        for r in range(2, self.last_row + 1):
            text = str(self.ws[f"B{r}"].value or "")
            if "ì˜¤ì¼€ì´ë§ˆíŠ¸" in text:
                site_rows["OK"].append(r)
            elif "ì•„ì´ì˜ˆìŠ¤" in text:
                site_rows["IY"].append(r)
        return site_rows

    def copy_to_new_sheet(self, 
                         wb: Workbook, 
                         sheet_name: str, 
                         row_indices: List[int]) -> None:
        """ì§€ì •ëœ í–‰ë“¤ì„ ìƒˆ ì‹œíŠ¸ë¡œ ë³µì‚¬"""
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            
        new_ws = wb.create_sheet(sheet_name)
        
        # í—¤ë”ì™€ ì—´ ë„ˆë¹„ ë³µì‚¬
        for c in range(1, self.last_col + 1):
            new_ws.cell(row=1, column=c, 
                       value=self.ws.cell(row=1, column=c).value)
            new_ws.column_dimensions[get_column_letter(c)].width = self.col_widths[c - 1]
        
        # ë°ì´í„° ë³µì‚¬
        for idx, r in enumerate(row_indices, start=2):
            for c in range(1, self.last_col + 1):
                new_ws.cell(row=idx, column=c, 
                          value=self.ws.cell(row=r, column=c).value)
            new_ws[f"A{idx}"].value = "=ROW()-1"


if __name__ == "__main__":
    test_file = "/Users/smith/Documents/github/OKMart/sabangnet_API/files/test-[ê¸°ë³¸ì–‘ì‹]-í•©í¬ì¥ìš©.xlsx"
    ali_merge_packaging(test_file)
    print("ëª¨ë“  ì²˜ë¦¬ê°€ ì™„ë£Œë˜ì—ˆìŠµë‹ˆë‹¤!")