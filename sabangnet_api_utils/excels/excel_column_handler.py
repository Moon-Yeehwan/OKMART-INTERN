from openpyxl.styles import Font, PatternFill, Alignment
import re


class ExcelColumnHandler:
    def __init__(self):
        self.red_font = Font(color="FF0000")

    def a_formula_column(self, cell):
        """
        순번 컬럼
        args:
            cell: 셀
        """
        cell.number_format = 'General'
        cell.value = "=ROW()-1"

    def a_value_column(self, cell):
        """
        순번 컬럼
        args:
            cell: 셀
        """
        cell.number_format = 'General'
        cell.value = cell.row - 1

    def d_column(self, cell, *source_cells):
        """
        금액 컬럼
        args:
            cell: 셀
            source_cells: 셀 값 / =O2+P2+V2 각 셀값 
                예 ) ws[f"O{row}"], ws[f"P{row}"], ws[f"V{row}"]
        """
        if cell.value is None:
            return

        cell.number_format = 'General'
        cell.value = str(sum(
            int(source_cell.value)
            for source_cell in source_cells
            if source_cell.value is not None and self._is_number(source_cell.value)
        ))

    def e_column(self, cell):
        """
        E 컬럼 포맷팅 - 문자타입으로 저장
        args:
            cell: 대상 셀
        """
        if cell.value and str(cell.value).replace('.', '').isdigit():
            num_str = str(cell.value).replace('.', '')
            cell.value = str(num_str)
            cell.number_format = '@'  # 텍스트 형식
        cell.alignment = Alignment(horizontal='right')

    def f_column(self, cell):
        """
        F 컬럼 포맷팅 - 텍스트 정리 및 조건부 하이라이팅
        args:
            cell: 대상 셀
        """
        # 텍스트 정리
        if cell.value:
            cell.value = str(cell.value).replace(' 1개', '')

        # 하이라이팅 조건 확인 및 적용
        if cell.value is not None and self._should_highlight_cell(cell.value):
            cell.fill = PatternFill(
                start_color="ADD8E6",
                end_color="ADD8E6",
                fill_type="solid"
            )

    def l_column(self, cell):
        """
        L 컬럼 포맷팅 - 신용, 착불 처리
        args:
            cell: 대상 셀
        """
        l_value_str = str(cell.value).strip()
        if l_value_str == "신용":
            cell.value = ""
        elif l_value_str == "착불":
            cell.font = self.red_font

    def h_i_column(self, cell):
        """
        H, I 컬럼 포맷팅 - 전화번호 포맷팅
        args:
            cell: 대상 셀
        """
        if cell.value:
            val = str(cell.value).replace('-', '').strip()
            if len(val) == 11 and val.startswith('010') and val.isdigit():
                cell.value = f"{val[:3]}-{val[3:7]}-{val[7:]}"

    def convert_int_column(self, cell):
        """
        셀의 문자열 숫자를 정수로 변환
        args:
            cell: 대상 셀
            row: 행
        """
        if cell.value is None:
            return
        if isinstance(cell.value, str):
            raw = cell.value.strip()
            # 숫자(0-9), 쉼표, 마침표만 허용
            if re.fullmatch(r"[0-9,\.]+", raw):
                cleaned = re.sub(r"[^\d]", "", raw)
                # 15자리 초과시 문자열로 저장
                if len(cleaned) >= 16:
                    cell.value = str(cell.value)
                    cell.number_format = '@'
                    return
                # 0 도 유효 숫자로 인정
                if raw not in {"", ".", ","}:
                    cell.value = int(cleaned) if cleaned else 0
                    cell.number_format = "0"

    def _convert_to_number(self, cell_value):
        """
        문자열을 숫자로 변환
        args:
            cell_value: 문자열 숫자
        """
        return float(cell_value) if '.' in str(cell_value) else int(float(cell_value))

    def _should_highlight_cell(self, txt: str) -> bool:
        """
        셀 값이 다음 조건 중 하나라도 만족하면 True:
        - 빈 문자열
        - 'none' (대소문자 무관)
        - 모든 문자가 '#'
        - 순수 숫자
        - '숫자개' 패턴 (예: '3개')
        """
        if not txt or txt.strip() == "":
            return True
        txt = txt.strip()
        if txt.lower() == "none":
            return True
        if all(c == '#' for c in txt):
            return True
        if txt.isdigit():
            return True
        if txt.endswith("개") and txt[:-1].isdigit():
            return True
        return False

    def _is_number(self, value):
        try:
            float(value)
            return True
        except (TypeError, ValueError):
            return False
