"""지그재그 합포장 자동화 모듈"""

from __future__ import annotations
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List

from openpyxl.styles import Font, PatternFill, Border
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from utils.excels.excel_handler import ExcelHandler


# 설정 상수
MALL_NAME = "지그재그"
BLUE_FILL = PatternFill(start_color="CCE8FF", end_color="CCE8FF", fill_type="solid")

# 시트 분리 설정 
ACCOUNT_MAPPING = {
    "OK": ["오케이마트"],
    "IY": ["아이예스"]
}

# 필수 생성 시트 목록
REQUIRED_SHEETS = list(ACCOUNT_MAPPING.keys())

FONT_MALGUN = Font(name="맑은 고딕", size=9)
HDR_FILL = PatternFill(start_color="006100", end_color="006100", fill_type="solid")
NO_BORDER = Border()

STAR_QTY_RE = re.compile(r"\* ?(\d+)")
MULTI_SEP_RE = re.compile(r"[\/;]")


def to_num(val) -> float:
    try:
        return float(re.sub(r"[^\d.-]", "", str(val))) if str(val).strip() else 0.0
    except ValueError:
        return 0.0


class ZIGZAGDataCleanerUtils:
    @staticmethod
    def clean_product_text(txt: str | None) -> str:
        """
        🔄 ExcelHandler 후보
        상품명 문자열 정리 - ' 1개' 제거
        """
        return str(txt).replace(" 1개", "").strip() if txt else ""

    @staticmethod
    def build_lookup_map(ws_lookup: Worksheet) -> Dict[str, str]:
        """
        Sheet1의 A:B를 딕셔너리로 변환
        (M열 → V열 매핑용 VLOOKUP 대체)
        """
        return {
            str(r[0]): r[1]
            for r in ws_lookup.iter_rows(min_row=2, max_col=2, values_only=True)
            if r[0] is not None
        }


def convert_m_column_to_int(ws: Worksheet) -> None:
    """
    🔄 ExcelHandler 후보
    M열 값을 정수로 변환
    """
    for row in range(2, ws.max_row + 1):
        try:
            cell = ws[f"M{row}"]
            cell.value = int(float(cell.value or 0))
        except (ValueError, TypeError):
            cell.value = 0


def highlight_multiple_items(ws: Worksheet) -> None:
    """
    🔄 ExcelHandler 후보
    F열에서 다중 수량 항목 파란색 배경으로 강조
    """
    for row in range(2, ws.max_row + 1):
        f_cell = ws[f"F{row}"]
        clean_text = ZIGZAGDataCleanerUtils.clean_product_text(f_cell.value)
        f_cell.value = clean_text
        
        # '개' 문자가 2회 이상 등장하면 파란색 배경
        if clean_text.count("개") >= 2:
            f_cell.fill = BLUE_FILL


class ZIGZAGSheetSplitter:
    """시트 분리 처리 클래스"""
    
    def __init__(self, ws: Worksheet):
        self.ws = ws
        self.last_row = ws.max_row
        self.last_col = ws.max_column
        
        # 열 너비 저장
        self.col_widths = [
            ws.column_dimensions[get_column_letter(c)].width
            for c in range(1, self.last_col + 1)
        ]

    def get_rows_by_sheet(self) -> Dict[str, List[int]]:
        """사이트별 행 번호 매핑 생성"""
        site_rows = defaultdict(list)
        for r in range(2, self.last_row + 1):
            text = str(self.ws[f"B{r}"].value or "")
            if "[오케이마트]" in text:
                site_rows["OK"].append(r)
            elif "[아이예스]" in text:
                site_rows["IY"].append(r)
        return site_rows

    def copy_to_new_sheet(self, 
                         wb: Workbook, 
                         sheet_name: str, 
                         row_indices: List[int]) -> None:
        """지정된 행들을 새 시트로 복사"""
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            
        new_ws = wb.create_sheet(sheet_name)
        
        # 헤더와 열 너비 복사
        for c in range(1, self.last_col + 1):
            new_ws.cell(row=1, column=c, 
                       value=self.ws.cell(row=1, column=c).value)
            new_ws.column_dimensions[get_column_letter(c)].width = self.col_widths[c - 1]
        
        # 데이터 복사
        for idx, r in enumerate(row_indices, start=2):
            for c in range(1, self.last_col + 1):
                new_ws.cell(row=idx, column=c, 
                          value=self.ws.cell(row=r, column=c).value)
            new_ws[f"A{idx}"].value = "=ROW()-1"

    def create_automation_sheet(self, wb) -> None:
        """
        매크로가 적용된 전체 시트를 "자동화"라는 이름으로 맨 앞에 복사
        """
        # 기존 "자동화" 시트가 있으면 삭제
        if "자동화" in wb.sheetnames:
            del wb["자동화"]
        
        # 현재 워크시트를 복사하여 "자동화" 시트 생성
        automation_ws = wb.copy_worksheet(self.ws)
        automation_ws.title = "자동화"
        
        # "자동화" 시트를 맨 앞으로 이동
        wb.move_sheet(automation_ws, offset=-len(wb.sheetnames) + 1)

    def copy_to_new_sheet_simple(self, 
                                wb: Workbook, 
                                sheet_name: str, 
                                row_indices: List[int] = None) -> None:
        """
        VBA와 동일한 방식으로 새 시트 생성
        - 원본 시트의 처리된 데이터를 각 계정명별로 복사만 수행
        - 추가 자동화 로직 적용하지 않음
        """
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            
        new_ws = wb.create_sheet(sheet_name)
        
        # 헤더와 열 너비 복사
        for c in range(1, self.last_col + 1):
            new_ws.cell(row=1, column=c, 
                       value=self.ws.cell(row=1, column=c).value)
            new_ws.column_dimensions[get_column_letter(c)].width = self.col_widths[c - 1]
        
        # 데이터 복사 (행이 있는 경우만)
        if row_indices:
            for idx, r in enumerate(row_indices, start=2):
                for c in range(1, self.last_col + 1):
                    new_ws.cell(row=idx, column=c, 
                              value=self.ws.cell(row=r, column=c).value)
                new_ws[f"A{idx}"].value = "=ROW()-1"


def zigzag_merge_packaging(input_path: str) -> str:
    """지그재그 주문 합포장 자동화 처리"""
    # Excel 파일 로드
    ex = ExcelHandler.from_file(input_path)
    ws = ex.ws

    # ========== 자동화 로직 적용 ==========
    # 1. 기본 서식 적용
    ex.set_basic_format()
    
    # 2. M열 정수 변환
    convert_m_column_to_int(ws)
    
    # 3. M열 → V열 VLOOKUP 처리
    if "Sheet1" in ex.wb.sheetnames:
        lookup_map = ZIGZAGDataCleanerUtils.build_lookup_map(ex.wb["Sheet1"])
        for row in range(2, ws.max_row + 1):
            m_val = str(ws[f"M{row}"].value)
            ws[f"V{row}"].value = lookup_map.get(m_val, "")
    
    # 4. D열 수식 설정 (=U+V)
    ex.autofill_d_column(formula="=U{row}+V{row}")
    
    # 5. 상품정보 처리 (다중수량 강조)
    highlight_multiple_items(ws)
    
    # 6. A열 순번 설정
    ex.set_row_number(ws)
    
    # 7. 열 정렬
    ex.set_column_alignment()
    
    # 8. 배경·테두리 제거
    ex.clear_fills_from_second_row()
    ex.clear_borders()
    
    # 9. C→B 정렬
    ex.sort_by_columns([2, 3])

    # ========== 자동화 시트 생성 (맨 앞에 위치) ==========
    # 매크로가 적용된 전체 시트를 "자동화" 이름으로 맨 앞에 복사
    splitter = ZIGZAGSheetSplitter(ws)
    splitter.create_automation_sheet(ex.wb)
    
    # ========== 시트분리 (OK, IY) ==========
    # 자동화 처리가 완료된 원본 시트에서 시트분리 수행
    rows_by_sheet = splitter.get_rows_by_sheet()
    
    # 모든 필수 시트 생성 (데이터 유무와 무관하게 OK, IY 시트 생성)
    for sheet_name in REQUIRED_SHEETS:
        splitter.copy_to_new_sheet_simple(
            ex.wb,
            sheet_name, 
            rows_by_sheet.get(sheet_name, [])
        )
    
    # 원본 시트 삭제 (자동화 시트로 대체되었으므로)
    original_sheet_name = ws.title
    if original_sheet_name in ex.wb.sheetnames and original_sheet_name != "자동화":
        del ex.wb[original_sheet_name]
    
    # 저장
    base_name = Path(input_path).stem  # 확장자 제거한 파일명
    output_path = ex.happojang_save_file(base_name=base_name)
    ex.wb.close()
    
    return output_path


if __name__ == "__main__":
    test_xlsx = "/Users/smith/Documents/github/OKMart/sabangnet_API/files/test-[기본양식]-합포장용.xlsx"
    zigzag_merge_packaging(test_xlsx)
    print("모든 처리가 완료되었습니다!")