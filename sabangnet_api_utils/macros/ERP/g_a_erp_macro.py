from openpyxl.cell import Cell
from utils.excels.excel_handler import ExcelHandler
from utils.excels.excel_column_handler import ExcelColumnHandler


class ERPGmaAucMacro:
    def __init__(self, file_path: str):
        self.ex = ExcelHandler.from_file(file_path)
        self.file_path = file_path
        self.ws = self.ex.ws
        self.wb = self.ex.wb
        self.basket_dict = {}
        self.headers = None

    def gauc_erp_macro_run(self):
        col_h = ExcelColumnHandler()
        self.basket_set = set()

        # 장바구니 중복 값 0으로 초기화
        for row in range(2, self.ws.max_row + 1):
            self._add_basket_dict(
                self.ws[f"Q{row}"], self.ws[f"V{row}"])
            
        # 시트 설정
        sheets_name = ["OK,CL,BB", "IY"]
        site_to_sheet = {
            "오케이마트": "OK,CL,BB",
            "클로버프": "OK,CL,BB",
            "베이지베이글": "OK,CL,BB",
            "아이예스": "IY",
        }

        # 정렬 기준: 2번째 컬럼(B) → 3번째 컬럼(C) 순으로 정렬
        sort_columns = [2, 3, -5]
        print("시트별 정렬, 시트 분리 시작...")
        headers, data = self.ex.preprocess_and_update_ws(self.ws, sort_columns)

        # 배송비 적용
        for row in range(self.ws.max_row, 1, -1):
            self._shipping_costs_column(self.ws[f"Q{row}"], self.ws[f"V{row}"])

        self.ex.split_and_write_ws_by_site(
            wb=self.wb,
            headers=headers,
            data=data,
            sheets_name=sheets_name,
            site_to_sheet=site_to_sheet,
            site_col_idx=2,
        )
        print("시트별 정렬, 시트 분리 완료")

        print("시트별 서식, 디자인 적용 시작...")
        for ws in self.wb.worksheets:
            self.ex.set_header_style(ws)
            if ws.max_row <= 1:
                continue
            for row in range(2, ws.max_row + 1):
                if ws.title != "자동화":
                    col_h.a_value_column(ws[f"A{row}"])
                else:
                    col_h.a_formula_column(ws[f"A{row}"])
                col_h.d_column(
                    ws[f"D{row}"], ws[f"O{row}"], ws[f"P{row}"], ws[f"V{row}"])
                col_h.e_column(ws[f"E{row}"])
                col_h.f_column(ws[f"F{row}"])
                col_h.l_column(ws[f"L{row}"])
                col_h.convert_int_column(ws[f"P{row}"])
                col_h.convert_int_column(ws[f"R{row}"])
                col_h.convert_int_column(ws[f"S{row}"])
                col_h.convert_int_column(ws[f"V{row}"])

            print(f"[{ws.title}] 서식 및 디자인 적용 완료")

        output_path = self.ex.save_file(self.file_path)
        print(f"✓ G,옥 ERP 자동화 완료! 최종 파일: {output_path}")
        return output_path

    def _add_basket_dict(self, basket_cell: Cell, shipping_cell: Cell):
        """
        장바구니 중복 값 추가
        args:
            basket_cell: 장바구니 번호 셀
            shipping_cell: 배송비 셀
        """
        basket_no = str(basket_cell.value).strip() if basket_cell.value else ""

        if not basket_no:
            return

        if basket_no not in self.basket_dict:
            if shipping_cell.value != 0 and shipping_cell.value != "":
                self.basket_dict[basket_no] = shipping_cell.value
                shipping_cell.value = 0
        else:
            shipping_cell.value = 0

    def _shipping_costs_column(self, basket_cell: Cell, shipping_cell: Cell):
        """
        정렬된 데이터에서 장바구니 중복 값 중 첫 번째 값에 배송비 적용
        args:
            basket_cell: 장바구니 번호 셀
            shipping_cell: 배송비 셀
        """
        basket_no = str(basket_cell.value).strip() if basket_cell.value else ""

        if basket_no in self.basket_dict:
            shipping_cell.value = self.basket_dict[basket_no]
            del self.basket_dict[basket_no]
